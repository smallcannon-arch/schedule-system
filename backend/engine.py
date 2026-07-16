# -*- coding: utf-8 -*-
"""
排課引擎 v1.6（CP-SAT）
用法：python 排課引擎.py 配課表範本.xlsx 課表輸出.xlsx
需求：Python 3.10+、openpyxl、ortools
v1.1：H13 資源班overlay、S05 行政空堂集中、S07 跨場地移動、科目欄動態讀取
v1.2：本土語分組自動推導、自然2+1場地模式、H17 自訂時段限制（教師/年級/班級）
v1.3：兩階段排課——「導師自排」科目留空位並輸出工作單（科目池+□空位）；
      資源班班級之國數（及有overlay之科目）例外，由引擎排課與overlay綁定
v1.4：新增母版 v5 九分頁格式、強化獨立檢核與輸入資料驗證
v1.5：新增 OpenAI 軟規則規劃介面，硬規則與最終驗證仍由 CP-SAT 負責
v1.6：新增完整度指標、正式模式自動排導師課、週節數上限與求解品質資訊
"""
import os
import re
import sys
from collections import defaultdict
from openpyxl import load_workbook
from ortools.sat.python import cp_model
import schedule_policy

DAYS = ["一", "二", "三", "四", "五"]
PERIODS = [1, 2, 3, 4, 5, 6, 7]
MORNING = {1, 2, 3, 4}
PAIR_START = [(1, 2), (2, 3), (3, 4), (5, 6), (6, 7)]  # 連堂不跨午休
V5_REQUIRED = {"班級", "教師", "場地", "科目節數", "配課", "年段時段", "本土語分組"}
V6_REQUIRED = {"班級", "教師與配課", "場地", "科目節數", "年段時段", "本土語分組"}


class InfeasibleScheduleError(RuntimeError):
    """A failed solve with deterministic, user-facing diagnostic details."""

    def __init__(self, message, diagnostics=(), status="INFEASIBLE"):
        super().__init__(message)
        self.diagnostics = list(diagnostics)
        self.status = status


def _diagnostic(title, detail, action, view, confirmed=True):
    return {"title": title, "detail": detail, "action": action,
            "view": view, "confirmed": bool(confirmed)}


def diagnose_infeasibility(d, tasks, candidates, invalid_locks=(), status="INFEASIBLE"):
    """Find necessary-capacity failures without claiming a complete unsat core."""
    diagnostics = []
    slots = [(day, period) for day in DAYS for period in PERIODS]

    invalid_by_key = {
        (lock["class"], lock["subj"], lock["day"], lock["p"]): lock
        for lock in invalid_locks
    }
    for lock in d.get("locks", []):
        key = (lock["class"], lock["subj"], lock["day"], lock["p"])
        if candidates.get(key) is None:
            invalid_by_key[key] = lock
    for lock in invalid_by_key.values():
        diagnostics.append(_diagnostic(
            f"{lock['class']} {lock['subj']} 的固定節次不可使用",
            f"週{lock['day']}第{lock['p']}節同時受到年段、班級、教師、科目或場地限制。",
            "回到鎖課來源或不排課時間，解除其中一項硬限制。", "lim"))

    class_locks = defaultdict(list)
    teacher_locks = defaultdict(list)
    room_locks = defaultdict(list)
    for lock in d.get("locks", []):
        key = (lock["class"], lock["subj"])
        task = tasks.get(key)
        class_locks[(lock["class"], lock["day"], lock["p"])].append(lock["subj"])
        teacher = _text(lock.get("teacher")) or (task["t"] if task else "")
        if teacher:
            teacher_locks[(teacher, lock["day"], lock["p"])].append(
                f"{lock['class']} {lock['subj']}")
        room = task["room"] if task else "R00"
        if room != "R00":
            room_locks[(room, lock["day"], lock["p"])].append(
                f"{lock['class']} {lock['subj']}")

    for (code, day, period), subjects in class_locks.items():
        unique = list(dict.fromkeys(subjects))
        if len(unique) > 1:
            diagnostics.append(_diagnostic(
                f"{code} 同一節被鎖定多門課",
                f"週{day}第{period}節同時鎖定：{'、'.join(unique)}。",
                "回到本土語、資源班或固定課來源，只保留一門課。", "rules"))
    for (teacher, day, period), lessons in teacher_locks.items():
        unique = list(dict.fromkeys(lessons))
        if len(unique) > 1:
            diagnostics.append(_diagnostic(
                f"{teacher} 的固定課發生衝堂",
                f"週{day}第{period}節同時需要：{'、'.join(unique)}。",
                "調整其中一筆固定課、資源班綁課或本土語分組。", "rules"))
    for (room, day, period), lessons in room_locks.items():
        capacity = int(d.get("rooms", {}).get(room, 1))
        if len(lessons) > capacity:
            diagnostics.append(_diagnostic(
                f"場地 {d.get('room_names', {}).get(room, room)} 容量不足",
                f"週{day}第{period}節容量 {capacity}，固定課需求 {len(lessons)}。",
                "改用其他教室、提高正確容量，或調整固定節次。", "build"))

    class_tasks = defaultdict(list)
    teacher_tasks = defaultdict(list)
    room_tasks = defaultdict(list)
    for key, task in tasks.items():
        class_tasks[key[0]].append((key, task))
        if task["t"]:
            teacher_tasks[task["t"]].append((key, task))
        if task["room"] != "R00":
            room_tasks[task["room"]].append((key, task))

    def candidate_slots(keys):
        wanted = set(keys)
        return {(day, period) for (code, subject, day, period), value in candidates.items()
                if value is not None and (code, subject) in wanted}

    for code, rows in class_tasks.items():
        demand = sum(task["h"] for _, task in rows)
        available = candidate_slots(key for key, _ in rows)
        if demand > len(available):
            diagnostics.append(_diagnostic(
                f"{code} 可排節次不足",
                f"引擎需排入 {demand} 節，但套用各項限制後只剩 {len(available)} 個候選節次。",
                "減少班級或年段不排課時間，或改由導師自排部分課程。", "lim"))

    native_daily = defaultdict(int)
    for group in d.get("native_groups", []):
        for teacher in (group.get("t"), group.get("assistant")):
            if teacher:
                native_daily[(teacher, group.get("d"))] += 1
    hard_cap = schedule_policy.daily_hard_cap(d)
    for teacher, rows in teacher_tasks.items():
        demand = sum(task["h"] for _, task in rows)
        available = candidate_slots(key for key, _ in rows)
        daily_capacity = sum(min(
            sum(1 for slot in available if slot[0] == day),
            max(0, hard_cap - native_daily[(teacher, day)])) for day in DAYS)
        if demand > daily_capacity:
            diagnostics.append(_diagnostic(
                f"{teacher}的授課容量不足",
                f"需由引擎安排 {demand} 節；扣除不排課、本土語固定課及每日 {hard_cap} 節上限後，最多可排 {daily_capacity} 節。",
                "檢查教師配課節數與不排課時間，必要時更換授課教師。", "alloc"))

    for room, rows in room_tasks.items():
        demand = sum(task["h"] for _, task in rows)
        available = candidate_slots(key for key, _ in rows)
        capacity = int(d.get("rooms", {}).get(room, 1))
        total_capacity = sum(capacity for day, period in available
                             if (room, day, period) not in d.get("room_blocked", set()))
        if demand > total_capacity:
            diagnostics.append(_diagnostic(
                f"{d.get('room_names', {}).get(room, room)} 的可用容量不足",
                f"課程需要 {demand} 節，依目前開放節次與容量最多只能容納 {total_capacity} 節。",
                "調整場地封鎖、場地容量或課程指定教室。", "lim"))

    for (code, subject), task in tasks.items():
        if not task["info"]["block"].startswith("2連堂") or task["h"] != 2:
            continue
        available = candidate_slots([(code, subject)])
        pairs = [(day, first, second) for day in DAYS for first, second in PAIR_START
                 if (day, first) in available and (day, second) in available]
        if not pairs:
            diagnostics.append(_diagnostic(
                f"{code} {subject} 找不到連堂時段",
                "所有可排節次中沒有同一天、且不跨午休的連續兩節。",
                "放寬該班、教師或場地限制，或取消此科的連堂硬規則。", "rules"))

    if not diagnostics:
        if status == "UNKNOWN":
            diagnostics.append(_diagnostic(
                "求解時間內尚未找到可行解",
                "目前不能判定為規則矛盾；案件規模或硬規則組合可能需要更多搜尋時間。",
                "先檢查固定課與教師不排課時間，再重新執行排課。", "rules", False))
        else:
            diagnostics.extend([
                _diagnostic(
                    "優先檢查固定課與綁課",
                    "多筆固定課、資源班綁課或本土語課可能共同占用同一教師、班級或場地。",
                    "依序暫時解除最近新增的鎖定，再重新排課確認。", "rules", False),
                _diagnostic(
                    "再檢查教師可用時間",
                    "個別課程看似有空間，但多位教師共用少數節次時仍可能形成整體衝突。",
                    "減少限制最密集教師的不排課時段，或調整配課。", "lim", False),
            ])
    return diagnostics[:10]


def _default_rules():
    weights = {"S01": 4, "S02": 4, "S03": 2, "S04": 2, "S05": 2,
               "S06": 1, "S07": 1, "S08": 2, "S09": 2}
    return {rid: {"on": True, "w": weight} for rid, weight in weights.items()}


def _text(value):
    return str(value or "").strip()


def _whole_number(value):
    if isinstance(value, bool) or value in (None, ""):
        raise ValueError
    number = float(value)
    if not number.is_integer():
        raise ValueError
    return int(number)


def _list_values(value):
    if isinstance(value, (list, tuple, set)):
        source = value
    else:
        source = re.split(r"[、,，;；\s]+", _text(value))
    return list(dict.fromkeys(_text(item) for item in source if _text(item)))


def _resource_sources(item):
    return _list_values(item.get("sources")) or [_text(item.get("class"))]


def _resource_pull_subjects(item):
    return _list_values(item.get("pull_subjects")) or [_text(item.get("subj"))]


def _normalize_subject_name(value):
    subject = _text(value)
    compact = re.sub(r"[\s（）()]", "", subject)
    if compact in {"閩南語", "臺語", "台語", "臺灣台語", "台灣台語", "本土語文閩南語"}:
        return "本土語文"
    return subject


def _is_minnan_language(value):
    language = re.sub(r"[\s（）()]", "", _text(value))
    return language in {
        "", "本土語", "本土語文", "閩南語", "臺語", "台語", "臺灣台語", "台灣台語",
        "本土語文閩南語",
    }


def _minnan_group_sources(groups):
    return {
        code
        for group in groups or []
        if _is_minnan_language(group.get("lang", group.get("language")))
        for code in _list_values(group.get("sources", group.get("classes")))
    }


def excel_safe(value):
    """Prevent user/model text from becoming an active Excel formula."""
    if not isinstance(value, str):
        return value
    stripped = value.lstrip()
    return "'" + value if stripped.startswith(("=", "+", "-", "@")) else value


def load_data(path):
    wb = load_workbook(path, data_only=True)
    if V5_REQUIRED.issubset(set(wb.sheetnames)) or V6_REQUIRED.issubset(set(wb.sheetnames)):
        return _load_data_v5(wb)
    return _load_data_v4(wb)


def load_frontend_data(payload, limits=(), rules=()):
    """Convert the browser's cloud-draft schema into the CP-SAT engine schema."""
    if not isinstance(payload, dict):
        raise ValueError("系統案件資料格式不正確")

    notes = []
    classes, class_codes = [], set()
    for row in payload.get("classes") or []:
        code = _text(row.get("code"))
        try:
            grade = int(row.get("g", row.get("grade")))
        except (TypeError, ValueError):
            raise ValueError(f"班級「{code or '未命名'}」的年級必須是 1 到 6")
        if not code or code in class_codes:
            raise ValueError("班級代碼不可空白或重複")
        if grade not in range(1, 7):
            raise ValueError(f"班級「{code}」的年級必須是 1 到 6")
        class_codes.add(code)
        classes.append({"code": code, "grade": grade, "tutor": _text(row.get("tutor")),
                        "res": bool(row.get("res", False))})
    if not classes:
        raise ValueError("系統案件尚未建立班級")

    roster = {_text(name): _text(role) for name, role in (payload.get("roster") or {}).items()
              if _text(name)}
    for item in classes:
        if item["tutor"]:
            roster.setdefault(item["tutor"], "導師")
    if not roster:
        raise ValueError("系統案件尚未建立教師")

    rooms = {str(key): max(1, int(value or 1))
             for key, value in (payload.get("rooms") or {"R00": 99}).items()}
    rooms.setdefault("R00", 99)
    room_names = {key: ("原班教室" if key == "R00" else key) for key in rooms}

    subjects = {}
    for name, row in (payload.get("subjects") or {}).items():
        subject = _text(name)
        if not subject:
            continue
        values = list(row.get("hours") or [])
        hours = {}
        for grade in range(1, 7):
            try:
                hours[grade] = int(values[grade - 1] or 0) if grade <= len(values) else 0
            except (TypeError, ValueError):
                raise ValueError(f"科目「{subject}」{grade}年級節數必須是整數")
            if hours[grade] < 0:
                raise ValueError(f"科目「{subject}」{grade}年級節數不可為負數")
        room = _text(row.get("room")) or "R00"
        if room not in rooms:
            raise ValueError(f"科目「{subject}」引用不存在的場地：{room}")
        subjects[subject] = {
            "hours": hours, "role": "", "room": room, "spread": None,
            "block": _text(row.get("block")),
            "banned": {int(value) for value in (row.get("banned") or [])
                       if str(value).isdigit()},
            "pair_mode": _text(row.get("pairMode", row.get("pair_mode"))),
            "self_arrange": bool(row.get("self", row.get("self_arrange", False))),
        }
    if not subjects:
        raise ValueError("系統案件尚未建立科目")

    grade_slot = {}
    source_slots = payload.get("gslot") or {}
    for grade in range(1, 7):
        rows = source_slots.get(str(grade), source_slots.get(grade)) or []
        for day_index, day in enumerate(DAYS):
            values = rows[day_index] if day_index < len(rows) else []
            for period_index, period in enumerate(PERIODS):
                grade_slot[(grade, day, period)] = bool(
                    values[period_index] if period_index < len(values) else 0)
    missing_grades = sorted({item["grade"] for item in classes
                             if not any(grade_slot[(item["grade"], day, period)]
                                        for day in DAYS for period in PERIODS)})
    if missing_grades:
        raise ValueError("年段時段沒有可排節次：" + "、".join(map(str, missing_grades)) + "年級")

    rule_map = _default_rules()
    for row in rules or payload.get("rules") or []:
        if not isinstance(row, (list, tuple)) or not row:
            continue
        rule_id = _text(row[0])
        if not rule_id:
            continue
        parameter = _text(row[4] if len(row) > 4 else "")
        match = re.search(r"penalty=(\d+)", parameter)
        rule_map[rule_id] = {"on": len(row) < 6 or _text(row[5]) == "是",
                             "w": int(match.group(1)) if match else 0}

    teacher_limit, grade_limit, class_limit = set(), set(), set()
    limit_rows = list(limits or payload.get("limits") or [])
    limit_rows.extend([list(row) + ["不可排", ""] for row in (payload.get("derived") or [])])
    cn_grade = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6}
    for row in limit_rows:
        if not isinstance(row, (list, tuple)) or len(row) < 4 or _text(row[3]) != "不可排":
            continue
        target, raw_day, raw_period = _text(row[0]), _text(row[1]), _text(row[2])
        days = DAYS if raw_day == "每日" else [raw_day]
        periods = PERIODS if raw_period == "全部" else [int(raw_period)] if raw_period.isdigit() else []
        for day in days:
            for period in periods:
                if day not in DAYS or period not in PERIODS:
                    continue
                if target in class_codes:
                    class_limit.add((target, day, period))
                elif "年級" in target:
                    grade = cn_grade.get(target[:1], int(target[:1]) if target[:1].isdigit() else 0)
                    if grade in range(1, 7):
                        grade_limit.add((grade, day, period))
                elif target:
                    teacher_limit.add((target, day, period))

    assign = {}
    source_assign = payload.get("assign") or {}
    for item in classes:
        row = source_assign.get(item["code"]) or {}
        for subject, info in subjects.items():
            teacher = _text(row.get(subject))
            if not teacher and info["self_arrange"] and item["tutor"]:
                teacher = item["tutor"]
            if teacher:
                assign[(item["code"], subject)] = teacher
                roster.setdefault(teacher, "其他")

    assignment_modes = {}
    for code, row in (payload.get("assignmentModes") or {}).items():
        if code not in class_codes or not isinstance(row, dict):
            continue
        for subject, mode in row.items():
            if subject in subjects and _text(mode) in {"tutor", "engine"}:
                assignment_modes[(code, subject)] = _text(mode)

    room_override = {}
    for code, row in (payload.get("override") or {}).items():
        if code not in class_codes or not isinstance(row, dict):
            continue
        for subject, room in row.items():
            if subject in subjects and _text(room):
                if _text(room) not in rooms:
                    raise ValueError(f"{code} {subject}引用不存在的場地：{room}")
                room_override[(code, subject)] = _text(room)

    locks = []
    for row in payload.get("locks") or []:
        code, day = _text(row.get("c", row.get("class"))), _text(row.get("d", row.get("day")))
        subject = _normalize_subject_name(row.get("s", row.get("subj")))
        try:
            period = int(row.get("p"))
        except (TypeError, ValueError):
            continue
        if code in class_codes and subject in subjects and day in DAYS and period in PERIODS:
            locks.append({"class": code, "day": day, "p": period, "subj": subject,
                          "teacher": _text(row.get("teacher")) or None})

    overlay = []
    class_grade = {item["code"]: item["grade"] for item in classes}
    for row_index, row in enumerate(payload.get("resGroups") or []):
        subject, teacher = _text(row.get("subj")), _text(row.get("t"))
        sources = _list_values(row.get("sources")) or [_text(row.get("code"))]
        sources = list(dict.fromkeys(code for code in sources if code))
        pull_subjects = _list_values(row.get("pullSubjects")) or [subject]
        pull_subjects = list(dict.fromkeys(
            _normalize_subject_name(value) for value in pull_subjects
            if _normalize_subject_name(value)))
        group_name = _text(row.get("grp")) or "資源班分組"
        if not sources or any(code not in class_codes for code in sources):
            raise ValueError(f"{group_name}引用不存在的來源班級")
        if len({class_grade[code] for code in sources}) != 1:
            raise ValueError(f"{group_name}的來源班級必須屬於同一年級")
        if subject not in subjects:
            raise ValueError(f"{group_name}的授課科目不存在：{subject}")
        if not pull_subjects or any(value not in subjects for value in pull_subjects):
            raise ValueError(f"{group_name}包含不存在的原班抽離科目")
        if not teacher:
            raise ValueError(f"{group_name}尚未指定資源班教師")
        count = min(10, max(1, int(row.get("n") or 1)))
        mode = "fixed" if _text(row.get("scheduleMode")) == "fixed" else "auto"
        slots = []
        for slot in row.get("slots") or []:
            day = _text(slot.get("d", slot.get("day")))
            try:
                period = int(slot.get("p", slot.get("period")))
            except (TypeError, ValueError):
                continue
            if day in DAYS and period in PERIODS:
                slots.append((day, period))
        slots = list(dict.fromkeys(slots))
        if mode == "fixed" and len(slots) != count:
            raise ValueError(
                f"{group_name}固定時段 {len(slots)} 節，與每週節數 {count} 不一致")
        roster.setdefault(teacher, "資源班教師")
        group_id = _text(row.get("id")) or f"resource-{row_index + 1}"
        for session_index in range(count):
            day, period = slots[session_index] if mode == "fixed" else (None, None)
            overlay.append({
                "id": f"{group_id}-{session_index + 1}", "grp": group_name,
                "class": sources[0], "sources": sources,
                "subj": subject, "pull_subjects": pull_subjects, "t": teacher,
                "day": day, "p": period,
            })

    blocked = set()
    for row in payload.get("blocked") or []:
        if not isinstance(row, (list, tuple)) or len(row) < 3:
            continue
        room, day = _text(row[0]), _text(row[1])
        try:
            period = int(row[2])
        except (TypeError, ValueError):
            continue
        if room in rooms and day in DAYS and period in PERIODS:
            blocked.add((room, day, period))

    native_rows = payload.get("nativeGroups") or []
    native_band_rows = payload.get("nativeBands") or []
    native_flag = payload.get("nativeLockEnabled")
    native_locks_present = any(lock["subj"] == "本土語文" for lock in locks)
    native_lock_enabled = (native_flag if isinstance(native_flag, bool)
                           else bool(native_band_rows or native_rows or native_locks_present))
    if native_flag is False and native_locks_present:
        raise ValueError("本土語課鎖定未啟用，但案件仍含固定鎖定資料")

    native_groups, native_slots, native_staff_slots = [], {}, set()
    native_room_load = defaultdict(int)
    native_group_names = set()
    for row in native_band_rows if native_lock_enabled else []:
        try:
            grade = int(row.get("g", row.get("grade")))
            period = int(row.get("p", row.get("period")))
        except (TypeError, ValueError, AttributeError):
            raise ValueError("本土語共同時段的年級或節次格式不正確")
        day = _text(row.get("d", row.get("day")))
        if grade not in range(1, 7) or day not in DAYS or period not in PERIODS:
            raise ValueError(f"本土語共同時段無效：{grade}年級 週{day}第{period}節")
        if grade in native_slots:
            raise ValueError(f"{grade}年級必須且只能設定一個本土語共同時段")
        native_slots[grade] = (day, period)

    # 舊案件沒有 nativeBands 時，仍可由原本分組列的星期、節次遷移。
    if native_lock_enabled and not native_slots:
        for row in native_rows:
            try:
                grade = int(row.get("g", row.get("grade")))
                period = int(row.get("p", row.get("period")))
            except (TypeError, ValueError, AttributeError):
                raise ValueError("本土語分組的年級或節次格式不正確")
            day = _text(row.get("d", row.get("day")))
            if grade not in range(1, 7) or day not in DAYS or period not in PERIODS:
                raise ValueError(f"本土語分組時段無效：{grade}年級 週{day}第{period}節")
            if grade in native_slots and native_slots[grade] != (day, period):
                raise ValueError(f"{grade}年級本土語分組必須使用相同星期與節次")
            native_slots[grade] = (day, period)

    for row in native_rows if native_lock_enabled else []:
        try:
            grade = int(row.get("g", row.get("grade")))
        except (TypeError, ValueError, AttributeError):
            raise ValueError("本土語分組的年級格式不正確")
        slot = native_slots.get(grade)
        if not slot:
            raise ValueError(f"{grade}年級尚未設定本土語共同時段")
        day, period = slot
        row_day = _text(row.get("d", row.get("day")))
        try:
            row_period = int(row.get("p", row.get("period"))) if row.get("p", row.get("period")) not in (None, "") else period
        except (TypeError, ValueError):
            raise ValueError("本土語分組的節次格式不正確")
        if row_day and (row_day, row_period) != slot:
            raise ValueError(f"{grade}年級本土語分組未使用年級共同時段")
        language = _text(row.get("lang", row.get("language"))) or "本土語文"
        group_name = (_text(row.get("grp", row.get("group"))) or
                      f"{grade}年級{language}{len(native_groups) + 1}組")
        if group_name in native_group_names:
            raise ValueError(f"本土語分組名稱重複：{group_name}")
        native_group_names.add(group_name)
        has_source_field = "sources" in row or "classes" in row
        sources = _list_values(row.get("sources", row.get("classes")))
        if not has_source_field:
            sources = [item["code"] for item in classes if item["grade"] == grade]
        if not sources:
            raise ValueError(f"{group_name}尚未填寫來源班級")
        for code in sources:
            source_class = next((item for item in classes if item["code"] == code), None)
            if not source_class:
                raise ValueError(f"{group_name}引用不存在的來源班級：{code}")
            if source_class["grade"] != grade:
                raise ValueError(f"{group_name}的來源班級 {code} 不屬於{grade}年級")
        try:
            students = max(0, int(row.get("students") or 0))
        except (TypeError, ValueError):
            raise ValueError(f"{group_name}的學生人數必須是整數")
        mode = _text(row.get("mode")) or ("直播共學" if "直播" in language else "實體")
        teacher = _text(row.get("t", row.get("teacher")))
        assistant = _text(row.get("assistant"))
        room = _text(row.get("room")) or "R00"
        if room not in rooms:
            raise ValueError(f"本土語分組引用不存在的場地：{room}")
        if not teacher:
            raise ValueError(f"{grade}年級本土語分組尚未填寫授課教師")
        if teacher not in roster:
            raise ValueError(f"{group_name}的授課教師不在教師名冊：{teacher}")
        if assistant and assistant not in roster:
            raise ValueError(f"{group_name}的協同教師不在教師名冊：{assistant}")
        if room != "R00":
            room_slot = (room, day, period)
            native_room_load[room_slot] += 1
            if native_room_load[room_slot] > rooms[room]:
                raise ValueError(f"{room}在週{day}第{period}節超過本土語分組可用容量")
        native_groups.append({"g": grade, "d": day, "p": period, "lang": language,
                              "grp": group_name, "sources": sources, "students": students,
                              "mode": mode, "t": teacher, "room": room, "assistant": assistant})
        if room != "R00":
            blocked.add((room, day, period))
        for name in (teacher, assistant):
            if name:
                staff_slot = (name, day, period)
                if staff_slot in native_staff_slots:
                    raise ValueError(f"{name}在週{day}第{period}節被重複指派本土語分組")
                native_staff_slots.add(staff_slot)
                teacher_limit.add((name, day, period))

    if native_lock_enabled and "本土語文" not in subjects:
        raise ValueError("已設定本土語分組，但科目節數缺少「本土語文」")
    if native_slots:
        for item in classes:
            slot = native_slots.get(item["grade"])
            if not slot:
                continue
            matching = [lock for lock in locks
                        if lock["class"] == item["code"] and lock["subj"] == "本土語文"]
            if not matching:
                locks.append({"class": item["code"], "day": slot[0], "p": slot[1],
                              "subj": "本土語文", "teacher": None})
            elif len(matching) != 1 or (matching[0]["day"], matching[0]["p"]) != slot:
                raise ValueError(f"{item['code']} 本土語固定節次與分組設定不一致")

    native_subject = subjects.get("本土語文")
    if native_subject and native_lock_enabled:
        slots_by_grade = defaultdict(set)
        for item in classes:
            hours = native_subject["hours"][item["grade"]]
            if not hours:
                continue
            if hours != 1:
                raise ValueError(f"{item['grade']}年級本土語文每週節數必須為 1")
            if item["grade"] not in native_slots:
                raise ValueError(f"{item['grade']}年級尚未建立本土語課鎖定分組")
            matching = [lock for lock in locks
                        if lock["class"] == item["code"] and lock["subj"] == "本土語文"]
            if len(matching) != 1:
                raise ValueError(f"{item['code']} 本土語文必須設定且只能有一個固定節次")
            if not grade_slot[(item["grade"], matching[0]["day"], matching[0]["p"])]:
                raise ValueError(f"{item['code']} 本土語固定節次不在該年級可排時段內")
            slots_by_grade[item["grade"]].add((matching[0]["day"], matching[0]["p"]))
        inconsistent = [grade for grade, slots in slots_by_grade.items() if len(slots) != 1]
        if inconsistent:
            raise ValueError("本土語分組必須全年級使用相同星期與節次：" +
                             "、".join(f"{grade}年級" for grade in inconsistent))

    for item in classes:
        required = sum(info["hours"][item["grade"]] for info in subjects.values())
        available = sum(grade_slot[(item["grade"], day, period)]
                        for day in DAYS for period in PERIODS)
        if required > available:
            raise ValueError(f"{item['code']} 每週課程 {required} 節，超過可排時段 {available} 節")

    policy_result = schedule_policy.validate_case(payload)
    teacher_caps = payload.get("tcap") or {}
    load = defaultdict(int)
    minnan_group_sources = _minnan_group_sources(native_groups)
    for (code, subject), teacher in assign.items():
        if native_lock_enabled and subject == "本土語文" and code in minnan_group_sources:
            continue
        grade = next(item["grade"] for item in classes if item["code"] == code)
        load[teacher] += subjects[subject]["hours"][grade]
    if native_lock_enabled:
        for group in native_groups:
            for teacher in (group.get("t"), group.get("assistant")):
                if teacher:
                    load[teacher] += 1
    for item in overlay:
        load[item["t"]] += 1
    teacher_weekly_quota, weekly_cap_violations = {}, []
    for teacher in roster:
        total = load.get(teacher, 0)
        if not schedule_policy.has_weekly_target(roster.get(teacher, "")):
            continue
        quota = schedule_policy.teacher_target(payload, teacher)
        if quota <= 0:
            continue
        teacher_weekly_quota[teacher] = quota
        if total != quota:
            message = f"教師應授節數不符：{teacher} 已配 {total} 節／應授 {quota} 節"
            notes.append(message)
            weekly_cap_violations.append(message)

    return {
        "schema_version": "frontend-1", "classes": classes, "roster": roster,
        "rooms": rooms, "room_names": room_names, "room_prio": {},
        "subjects": subjects, "grade_slot": grade_slot, "rules": rule_map,
        "teacher_limit": teacher_limit, "grade_limit": grade_limit,
        "class_limit": class_limit, "assign": assign, "assignment_modes": assignment_modes,
        "room_override": room_override,
        "locks": locks, "prefs": [], "overlay": overlay, "room_blocked": blocked,
        "native_bands": [{"g": grade, "d": slot[0], "p": slot[1]}
                         for grade, slot in sorted(native_slots.items())],
        "native_groups": native_groups, "native_lock_enabled": native_lock_enabled,
        "teacher_weekly_quota": teacher_weekly_quota,
        "teacher_weekly_load": dict(load),
        "weekly_cap_violations": weekly_cap_violations, "derived_notes": notes,
        "policy": schedule_policy.metadata(payload),
        "compliance_blocking_issues": policy_result["blocking"],
        "compliance_warnings": policy_result["warnings"],
    }


def _load_data_v5(wb):
    is_v6 = "教師與配課" in wb.sheetnames
    required = V6_REQUIRED if is_v6 else V5_REQUIRED
    missing = sorted(required - set(wb.sheetnames))
    if missing:
        raise ValueError(f"母版 v{6 if is_v6 else 5} 缺少工作表：" + "、".join(missing))

    d = {"schema_version": 6 if is_v6 else 5}
    notes = []

    # 班級
    classes, seen_classes = [], set()
    for row_number, row in enumerate(
            wb["班級"].iter_rows(min_row=2, values_only=True), start=2):
        code = _text(row[0] if len(row) > 0 else None)
        if not code:
            if any(_text(value) for value in row):
                raise ValueError(f"班級工作表第 {row_number} 列缺少班級代碼")
            continue
        try:
            grade = _whole_number(row[1] if len(row) > 1 else None)
        except (TypeError, ValueError):
            raise ValueError(f"班級工作表第 {row_number} 列：{code}的年級必須是 1 到 6")
        if not 1 <= grade <= 6:
            raise ValueError(f"班級「{code}」的年級超出 1 到 6")
        if code in seen_classes:
            raise ValueError(f"班級代碼重複：{code}")
        seen_classes.add(code)
        classes.append({"code": code, "grade": grade, "tutor": _text(row[2]),
                        "res": _text(row[3]) == "是"})
    if not classes:
        raise ValueError("「班級」工作表沒有有效班級")
    d["classes"] = classes

    # 教師：v6 讀取合併工作表左側；v5 仍支援獨立教師工作表。
    roster = {c["tutor"]: "導師" for c in classes if c["tutor"]}
    teacher_caps = {}
    teacher_rows = {}
    ws = wb["教師與配課"] if is_v6 else wb["教師"]
    header_row = 2 if is_v6 else 1
    first_data_row = header_row + 1
    headers = [_text(c.value).replace(" ▾", "") for c in ws[header_row]][:7]
    cap_col = next((i for i, h in enumerate(headers) if "上限" in h), 2)
    minus_col = next((i for i, h in enumerate(headers) if "減課" in h), 3)
    for row_number, row in enumerate(
            ws.iter_rows(min_row=first_data_row, max_col=7, values_only=True),
            start=first_data_row):
        name = _text(row[0] if row else None)
        if not name:
            if any(_text(value) for value in (row or ())):
                raise ValueError(f"教師工作表第 {row_number} 列缺少教師姓名")
            continue
        role = _text(row[1] if len(row) > 1 else None) or roster.get(name, "其他")
        cap = row[cap_col] if cap_col < len(row) else None
        minus = row[minus_col] if minus_col < len(row) else None
        try:
            normalized_cap = 0 if cap in (None, "") else _whole_number(cap)
            normalized_minus = 0 if minus in (None, "") else _whole_number(minus)
        except (TypeError, ValueError):
            raise ValueError(f"教師工作表第 {row_number} 列的節數上限與減課節數必須是整數")
        if normalized_cap < 0 or normalized_minus < 0:
            raise ValueError(f"教師工作表第 {row_number} 列的節數上限與減課節數不可為負數")
        incoming = {"role": role, "cap": normalized_cap, "minus": normalized_minus,
                    "row": row_number}
        previous = teacher_rows.get(name)
        if previous:
            conflicts = [label for key, label in (("role", "身分"), ("cap", "每週節數上限"),
                                                   ("minus", "減課節數"))
                         if previous[key] != incoming[key]]
            if conflicts:
                raise ValueError(
                    f"教師工作表第 {row_number} 列：{name}與第 {previous['row']} 列重複，且"
                    + "、".join(conflicts) + "不一致")
            notes.append(f"教師工作表第 {row_number} 列：{name}重複，已合併為同一位教師")
            continue
        teacher_rows[name] = incoming
        roster[name] = role
        teacher_caps[name] = {"cap": normalized_cap, "minus": normalized_minus}
    d["roster"], d["teacher_caps"] = roster, teacher_caps

    # 場地：v5 以場地名稱為識別；原班教室統一使用 R00。
    rooms, room_names, room_prio = {"R00": 999}, {"R00": "原班教室"}, {}

    def room_key(value):
        name = _text(value)
        return "R00" if not name or "原班" in name else name

    for row in wb["場地"].iter_rows(min_row=2, values_only=True):
        name = _text(row[0] if row else None)
        if not name:
            continue
        rid = room_key(name)
        try:
            capacity = int(row[1] or 1)
        except (TypeError, ValueError):
            raise ValueError(f"場地「{name}」容量必須是正整數")
        if capacity < 1:
            raise ValueError(f"場地「{name}」容量必須大於 0")
        rooms[rid], room_names[rid] = capacity, name
        if len(row) > 2 and row[2] not in (None, "") and rid != "R00":
            room_prio[rid] = int(row[2])
    d["rooms"], d["room_names"], d["room_prio"] = rooms, room_names, room_prio

    # 科目節數
    subjects = {}
    for row_number, row in enumerate(
            wb["科目節數"].iter_rows(min_row=2, values_only=True), start=2):
        name = _text(row[0] if row else None)
        if not name:
            if row and any(_text(value) for value in row):
                raise ValueError(f"科目節數第 {row_number} 列缺少科目名稱")
            continue
        if name in subjects:
            raise ValueError(f"科目重複：{name}")
        hours = {}
        for grade in range(1, 7):
            raw = row[grade] if grade < len(row) else 0
            try:
                hours[grade] = 0 if raw in (None, "") else _whole_number(raw)
            except (TypeError, ValueError):
                raise ValueError(f"科目「{name}」{grade}年級節數必須是整數")
            if hours[grade] < 0:
                raise ValueError(f"科目「{name}」{grade}年級節數不可為負數")
        special = _text(row[9] if len(row) > 9 else None)
        subjects[name] = {
            "hours": hours,
            "role": "",
            "room": room_key(row[7] if len(row) > 7 else None),
            "spread": None,
            "block": "2+1" if special == "2+1分兩天" else "2連堂" if special == "兩節連堂" else "",
            "banned": {4, 5} if "不排第4" in special else set(),
            "pair_mode": "全程指定場地",
            "self_arrange": _text(row[8] if len(row) > 8 else None) == "是",
        }
        if subjects[name]["room"] not in rooms:
            raise ValueError(f"科目「{name}」引用不存在的場地：{subjects[name]['room']}")
    if not subjects:
        raise ValueError("「科目節數」工作表沒有有效科目")
    d["subjects"], d["rules"] = subjects, _default_rules()

    # 年段時段
    grade_slot = {}
    grades_seen = set()
    for row_number, row in enumerate(
            wb["年段時段"].iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] in (None, ""):
            if row and any(_text(value) for value in row[1:]):
                raise ValueError(f"年段時段第 {row_number} 列缺少年級")
            continue
        try:
            grade = _whole_number(row[0])
        except (TypeError, ValueError):
            raise ValueError(f"年段時段第 {row_number} 列的年級必須是 1 到 6")
        if not 1 <= grade <= 6:
            raise ValueError(f"年段時段第 {row_number} 列的年級必須是 1 到 6")
        if grade in grades_seen:
            raise ValueError(f"年段時段第 {row_number} 列：{grade}年級重複設定")
        grades_seen.add(grade)
        for day_i, day in enumerate(DAYS):
            for p_i, period in enumerate(PERIODS):
                col = 1 + day_i * 7 + p_i
                grade_slot[(grade, day, period)] = bool(row[col] if col < len(row) else 0)
    missing_grades = sorted({c["grade"] for c in classes} - grades_seen)
    if missing_grades:
        raise ValueError("年段時段缺少年級：" + "、".join(map(str, missing_grades)))
    d["grade_slot"] = grade_slot

    # 本土語分組：每個年級建立一次固定課鎖定；場地、授課教師與協同教師同步封鎖。
    locks, blocked, teacher_limit, band_seen = [], set(), set(), set()
    native_groups, native_group_names, native_room_load = [], set(), defaultdict(int)
    class_by_code = {item["code"]: item for item in classes}
    native_ws = wb["本土語分組"]
    native_headers = [_text(cell.value).replace(" ▾", "") for cell in native_ws[1]]

    def native_col(label, fallback):
        return next((i for i, header in enumerate(native_headers) if label in header), fallback)

    native_idx = {
        "grade": native_col("年級", 0), "day": native_col("星期", 1),
        "period": native_col("節次", 2), "language": native_col("語別", 3),
        "teacher": native_col("授課教師", 4), "room": native_col("教室", 5),
        "assistant": native_col("協同教師", 6), "group": native_col("分組名稱", 7),
        "sources": native_col("來源班級", 8), "students": native_col("學生人數", 9),
        "mode": native_col("上課方式", 10),
    }
    for row_number, row in enumerate(
            native_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(_text(value) for value in row):
            continue
        try:
            grade = _whole_number(row[native_idx["grade"]])
            period = _whole_number(row[native_idx["period"]])
        except (TypeError, ValueError):
            raise ValueError(f"本土語分組第 {row_number} 列的年級或節次無效")
        day = _text(row[native_idx["day"]])
        if grade not in range(1, 7) or day not in DAYS or period not in PERIODS:
            raise ValueError(
                f"本土語分組第 {row_number} 列時段無效：{grade}年級 週{day} 第{period}節")
        previous = band_seen and next((x for x in band_seen if x[0] == grade), None)
        if previous and previous[1:] != (day, period):
            raise ValueError(f"本土語分組同一年級出現不同時段：{grade}年級")
        if not previous:
            band_seen.add((grade, day, period))
            for c in classes:
                if c["grade"] == grade:
                    locks.append({"class": c["code"], "day": day, "p": period,
                                  "subj": "本土語文", "teacher": None})
        rid = room_key(row[native_idx["room"]] if len(row) > native_idx["room"] else None)
        if rid not in rooms:
            raise ValueError(f"本土語分組引用不存在的場地：{rid}")
        if rid != "R00":
            room_slot = (rid, day, period)
            native_room_load[room_slot] += 1
            if native_room_load[room_slot] > rooms[rid]:
                raise ValueError(f"{rid}在週{day}第{period}節超過本土語分組可用容量")
            blocked.add((rid, day, period))
        language = _text(row[native_idx["language"]] if len(row) > native_idx["language"] else None) or "本土語文"
        teacher = _text(row[native_idx["teacher"]] if len(row) > native_idx["teacher"] else None)
        assistant = _text(row[native_idx["assistant"]] if len(row) > native_idx["assistant"] else None)
        group_name = _text(row[native_idx["group"]] if len(row) > native_idx["group"] else None) or f"{grade}年級{language}{len(native_groups) + 1}組"
        if group_name in native_group_names:
            raise ValueError(f"本土語分組名稱重複：{group_name}")
        native_group_names.add(group_name)
        sources = _list_values(row[native_idx["sources"]] if len(row) > native_idx["sources"] else None) or [
            item["code"] for item in classes if item["grade"] == grade]
        for code in sources:
            if code not in class_by_code:
                raise ValueError(f"{group_name}引用不存在的來源班級：{code}")
            if class_by_code[code]["grade"] != grade:
                raise ValueError(f"{group_name}的來源班級 {code} 不屬於{grade}年級")
        try:
            raw_students = row[native_idx["students"]] if len(row) > native_idx["students"] else None
            students = 0 if raw_students in (None, "") else _whole_number(raw_students)
        except (TypeError, ValueError):
            raise ValueError(f"{group_name}的學生人數必須是整數")
        if students < 0:
            raise ValueError(f"{group_name}的學生人數不可為負數")
        mode = _text(row[native_idx["mode"]] if len(row) > native_idx["mode"] else None) or ("直播共學" if "直播" in language else "實體")
        if not teacher:
            raise ValueError(f"{group_name}尚未填寫授課教師")
        for name, label in ((teacher, "授課教師"), (assistant, "協同教師")):
            if not name:
                continue
            if name not in roster:
                raise ValueError(f"{group_name}的{label}不在教師名冊：{name}")
            staff_slot = (name, day, period)
            if staff_slot in teacher_limit:
                raise ValueError(f"{name}在週{day}第{period}節被重複指派本土語分組")
            teacher_limit.add(staff_slot)
        native_groups.append({"g": grade, "d": day, "p": period, "lang": language,
                              "grp": group_name, "sources": sources, "students": students,
                              "mode": mode, "t": teacher, "room": rid, "assistant": assistant})
    if locks and "本土語文" not in subjects:
        raise ValueError("本土語分組有資料，但科目節數缺少「本土語文」")
    d["locks"], d["room_blocked"] = locks, blocked
    d["native_bands"] = [{"g": grade, "d": day, "p": period}
                         for grade, day, period in sorted(band_seen)]
    d["native_groups"] = native_groups
    d["teacher_limit"], d["grade_limit"], d["class_limit"] = teacher_limit, set(), set()

    # 配課：v6 讀取合併工作表右側 I:K；v5 讀取獨立配課工作表。
    assign = {}
    for c in classes:
        if not c["tutor"]:
            continue
        for subject, info in subjects.items():
            if info["self_arrange"] and info["hours"][c["grade"]] > 0:
                assign[(c["code"], subject)] = c["tutor"]
    class_map = {c["code"]: c for c in classes}
    cn_grade = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6}

    def expand_classes(spec, subject, row_number):
        result = []
        for token in re.split(r"[、,，;；\s]+", _text(spec)):
            if not token:
                continue
            if token in ("全部", "全校"):
                result.extend(c["code"] for c in classes if subjects[subject]["hours"][c["grade"]] > 0)
                continue
            match = re.fullmatch(r"([1-6一二三四五六])年級(?:全部)?", token)
            if match:
                grade = cn_grade.get(match.group(1), int(match.group(1)) if match.group(1).isdigit() else 0)
                result.extend(c["code"] for c in classes if c["grade"] == grade and subjects[subject]["hours"][grade] > 0)
            elif token in class_map:
                result.append(token)
            else:
                raise ValueError(
                    f"配課工作表第 {row_number} 列：{subject}的班級「{token}」不存在")
        return list(dict.fromkeys(result))

    assignment_ws = wb["教師與配課"] if is_v6 else wb["配課"]
    assignment_row = 3 if is_v6 else 2
    assignment_offset = 8 if is_v6 else 0
    explicit_assignments = {}
    for row_number, row in enumerate(
            assignment_ws.iter_rows(min_row=assignment_row, values_only=True),
            start=assignment_row):
        teacher = _text(row[assignment_offset] if len(row) > assignment_offset else None)
        subject = _normalize_subject_name(
            row[assignment_offset + 1] if len(row) > assignment_offset + 1 else None)
        if not teacher and not subject:
            target = _text(row[assignment_offset + 2] if len(row) > assignment_offset + 2 else None)
            if target:
                raise ValueError(f"配課工作表第 {row_number} 列缺少教師與科目")
            continue
        if not teacher or not subject:
            raise ValueError(f"配課工作表第 {row_number} 列必須完整填寫教師、科目與任教班級")
        if subject not in subjects:
            raise ValueError(f"配課工作表第 {row_number} 列：科目「{subject}」不在科目節數表")
        if teacher not in roster:
            raise ValueError(f"配課工作表第 {row_number} 列：教師「{teacher}」未列於教師表")
        raw_targets = row[assignment_offset + 2] if len(row) > assignment_offset + 2 else None
        if not _text(raw_targets):
            raise ValueError(f"配課工作表第 {row_number} 列缺少任教班級")
        targets = expand_classes(raw_targets, subject, row_number)
        if not targets:
            notes.append(
                f"配課工作表第 {row_number} 列：{teacher}／{subject}目前沒有符合的任教班級，已略過")
            continue
        for code in targets:
            if subjects[subject]["hours"][class_map[code]["grade"]] > 0:
                key = (code, subject)
                previous = explicit_assignments.get(key)
                if previous and previous["teacher"] != teacher:
                    raise ValueError(
                        f"配課工作表第 {row_number} 列：{code} {subject}已在第 {previous['row']} 列"
                        f"配給{previous['teacher']}")
                explicit_assignments[key] = {"teacher": teacher, "row": row_number}
                assign[(code, subject)] = teacher
    d["assign"] = assign
    d["room_override"], d["prefs"] = {}, []

    # v6 選用分頁：與瀏覽器匯入使用相同語意，避免 Excel 直傳漏掉硬規則。
    teacher_limit = set(d.get("teacher_limit") or set())
    grade_limit = set(d.get("grade_limit") or set())
    class_limit = set(d.get("class_limit") or set())
    limit_sheet = next((wb[name] for name in ("不排課時間", "教師時段限制")
                        if name in wb.sheetnames), None)
    if limit_sheet:
        cn_grade = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6}
        class_codes = {item["code"] for item in classes}
        for row_number, row in enumerate(
                limit_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(_text(value) for value in row):
                continue
            target = _text(row[0] if len(row) > 0 else None)
            raw_day = _text(row[1] if len(row) > 1 else None)
            raw_period = _text(row[2] if len(row) > 2 else None)
            kind = _text(row[3] if len(row) > 3 else None) or "不可排"
            if not target or not raw_day or not raw_period:
                raise ValueError(
                    f"{limit_sheet.title}第 {row_number} 列必須完整填寫對象、星期與節次")
            if kind != "不可排":
                continue
            if raw_day != "每日" and raw_day not in DAYS:
                raise ValueError(f"{limit_sheet.title}第 {row_number} 列的星期不正確：{raw_day}")
            if raw_period != "全部":
                try:
                    parsed_period = _whole_number(raw_period)
                except (TypeError, ValueError):
                    raise ValueError(
                        f"{limit_sheet.title}第 {row_number} 列的節次不正確：{raw_period}")
                if parsed_period not in PERIODS:
                    raise ValueError(
                        f"{limit_sheet.title}第 {row_number} 列的節次不正確：{raw_period}")
            days = DAYS if raw_day == "每日" else [raw_day]
            periods = PERIODS if raw_period == "全部" else [parsed_period]
            for day in days:
                for period in periods:
                    if target in class_codes:
                        class_limit.add((target, day, period))
                    elif "年級" in target:
                        grade = cn_grade.get(target[:1], int(target[:1]) if target[:1].isdigit() else 0)
                        if grade in range(1, 7):
                            grade_limit.add((grade, day, period))
                    else:
                        teacher_limit.add((target, day, period))
    d["teacher_limit"], d["grade_limit"], d["class_limit"] = (
        teacher_limit, grade_limit, class_limit)

    overlay = []
    if "資源班overlay" in wb.sheetnames:
        overlay_sheet = wb["資源班overlay"]
        for row_number, row in enumerate(
                overlay_sheet.iter_rows(min_row=2, values_only=True), start=2):
            group = _text(row[0] if len(row) > 0 else None) or f"第{row_number}列"
            code = _text(row[1] if len(row) > 1 else None)
            subject = _normalize_subject_name(row[2] if len(row) > 2 else None)
            teacher = _text(row[3] if len(row) > 3 else None)
            if not code and not subject and not teacher:
                continue
            if not code or not subject or not teacher:
                raise ValueError(
                    f"資源班overlay第 {row_number} 列必須完整填寫來源班級、科目與教師")
            if code not in class_map:
                raise ValueError(f"資源班overlay第 {row_number} 列的來源班級不存在：{code}")
            if subject not in subjects:
                raise ValueError(f"資源班overlay第 {row_number} 列的科目不存在：{subject}")
            if teacher not in roster:
                raise ValueError(f"資源班overlay第 {row_number} 列的教師不在名冊：{teacher}")
            day = _text(row[4] if len(row) > 4 else None)
            raw_period = row[5] if len(row) > 5 else None
            try:
                period = _whole_number(raw_period) if raw_period not in (None, "") else None
            except (TypeError, ValueError):
                raise ValueError(f"資源班overlay第 {row_number} 列的節次不正確：{group}")
            if day and day not in DAYS or period is not None and period not in PERIODS:
                raise ValueError(f"資源班overlay第 {row_number} 列的時段不正確：{group}")
            overlay.append({"grp": group, "class": code, "subj": subject, "t": teacher,
                            "day": day or None, "p": period})
    d["overlay"] = overlay

    # 基礎一致性檢核與教師週節數提示。
    for c in classes:
        need = sum(info["hours"][c["grade"]] for info in subjects.values())
        available = sum(grade_slot[(c["grade"], day, p)] for day in DAYS for p in PERIODS)
        if need > available:
            raise ValueError(f"{c['code']} 每週課程 {need} 節，超過該年級可排時段 {available} 節")
    d["policy"] = {"profileId": schedule_policy.PROFILE_ID}
    d["tcap"] = teacher_caps
    load = defaultdict(int)
    minnan_group_sources = _minnan_group_sources(native_groups)
    for (code, subject), teacher in assign.items():
        if native_groups and subject == "本土語文" and code in minnan_group_sources:
            continue
        load[teacher] += subjects[subject]["hours"][class_map[code]["grade"]]
    for group in native_groups:
        for teacher in (group.get("t"), group.get("assistant")):
            if teacher:
                load[teacher] += 1
    for item in overlay:
        load[item["t"]] += 1
    teacher_weekly_quota, weekly_cap_violations = {}, []
    for teacher in roster:
        total = load.get(teacher, 0)
        if not schedule_policy.has_weekly_target(roster.get(teacher, "")):
            continue
        quota = schedule_policy.teacher_target(d, teacher)
        if quota <= 0:
            continue
        teacher_weekly_quota[teacher] = quota
        if total != quota:
            message = f"教師應授節數不符：{teacher} 已配 {total} 節／應授 {quota} 節"
            notes.append(message)
            weekly_cap_violations.append(message)
    d["teacher_weekly_quota"] = teacher_weekly_quota
    d["teacher_weekly_load"] = dict(load)
    d["weekly_cap_violations"] = weekly_cap_violations
    policy_result = schedule_policy.validate_case(d)
    d["policy_meta"] = schedule_policy.metadata(d)
    d["compliance_blocking_issues"] = policy_result["blocking"]
    d["compliance_warnings"] = policy_result["warnings"]
    d["derived_notes"] = notes
    return d


def _load_data_v4(wb):
    d = {}

    ws = wb["設定_年段時段"]
    grade_slot = {}
    for r in range(2, 8):
        g = ws.cell(r, 1).value
        if not g:
            continue
        col = 2
        for day in DAYS:
            for p in PERIODS:
                grade_slot[(int(g), day, p)] = bool(ws.cell(r, col).value)
                col += 1
    d["grade_slot"] = grade_slot

    ws = wb["設定_場地表"]
    rooms, room_names, room_prio = {}, {}, {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if rid:
            rooms[rid] = int(ws.cell(r, 3).value or 1)
            room_names[rid] = str(ws.cell(r, 2).value or "")
            pr = ws.cell(r, 5).value
            if pr:
                room_prio[rid] = int(pr)
    d["rooms"], d["room_names"], d["room_prio"] = rooms, room_names, room_prio

    ws = wb["場地封鎖"]
    blocked = set()
    for r in range(2, ws.max_row + 1):
        rid, day, p = ws.cell(r, 1).value, ws.cell(r, 3).value, ws.cell(r, 4).value
        if rid and day and p:
            blocked.add((rid, str(day), int(p)))
    d["room_blocked"] = blocked

    ws = wb["設定_科目表"]
    subjects = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name or not ws.cell(r, 1).value:
            continue
        subjects[name] = {
            "hours": {g: int(ws.cell(r, 5 + g).value or 0) for g in range(1, 7)},
            "role": ws.cell(r, 12).value or "",
            "room": ws.cell(r, 13).value or "R00",
            "spread": ws.cell(r, 15).value,
            "block": str(ws.cell(r, 16).value or ""),
            "banned": {int(x) for x in str(ws.cell(r, 17).value or "").replace("，", ",").split(",") if x.strip().isdigit()},
            "pair_mode": str(ws.cell(r, 19).value or ""),
            "self_arrange": str(ws.cell(r, 20).value or "") == "導師自排",
        }
    d["subjects"] = subjects

    ws = wb["設定_排課規則清單"]
    rules = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if not rid:
            continue
        param = str(ws.cell(r, 5).value or "")
        w = int(param.split("=")[1]) if "penalty=" in param else 0
        rules[rid] = {"on": (ws.cell(r, 6).value == "是"), "w": w}
    d["rules"] = rules

    ws = wb["教師名冊"]
    roster = {}
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, 1).value
        if nm:
            roster[nm] = ws.cell(r, 2).value or ""
    d["roster"] = roster

    ws = wb["教師時段限制"]
    tlimit, glimit, climit = set(), set(), set()
    CN = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6}
    for r in range(2, ws.max_row + 1):
        nm, day, p, typ = (ws.cell(r, c).value for c in (1, 2, 3, 4))
        if not nm or typ != "不可排":
            continue
        nm = str(nm).strip()
        days = DAYS if day == "每日" else [str(day)]
        ps = PERIODS if str(p) == "全部" else [int(p)]
        for dd in days:
            for pp in ps:
                if "年級" in nm:  # 年級：如「3年級」「三年級」→ 學年共同時間
                    g = CN.get(nm[0], None)
                    g = g or (int(nm[0]) if nm[0].isdigit() else None)
                    if g:
                        glimit.add((g, dd, pp))
                elif any(ch in "甲乙丙丁戊己庚辛" for ch in nm):  # 班級代碼
                    climit.add((nm, dd, pp))
                else:  # 教師
                    tlimit.add((nm, dd, pp))
    d["teacher_limit"], d["grade_limit"], d["class_limit"] = tlimit, glimit, climit

    ws = wb["班級基本資料"]
    classes = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if code:
            classes.append({"code": str(code), "grade": int(ws.cell(r, 2).value),
                            "tutor": str(ws.cell(r, 3).value or ""),
                            "res": str(ws.cell(r, 4).value or "") == "是"})
    d["classes"] = classes

    ws = wb["配課表"]
    subj_names = []
    for c in range(4, 40):
        h = ws.cell(1, c).value
        if not h:
            break
        subj_names.append(h)
    assign = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code:
            continue
        for i, s in enumerate(subj_names):
            t = ws.cell(r, 4 + i).value
            if t:
                assign[(str(code), s)] = str(t)
    d["assign"] = assign

    ws = wb["場地例外(room_override)"]
    override = {}
    for r in range(2, ws.max_row + 1):
        code, s, rid = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if code and s and rid:
            override[(str(code), s)] = rid
    d["room_override"] = override

    ws = wb["固定課鎖定"]
    locks, prefs = [], []
    for r in range(2, ws.max_row + 1):
        code, day, p, s = (ws.cell(r, c).value for c in (1, 2, 3, 4))
        typ = ws.cell(r, 7).value
        if not (code and day and p and s):
            continue
        item = {"class": str(code), "day": str(day), "p": int(p), "subj": s,
                "teacher": ws.cell(r, 5).value}
        (locks if typ == "鎖定" else prefs).append(item)
    d["locks"], d["prefs"] = locks, prefs

    ws = wb["資源班overlay"]
    overlay = []  # H13：一列=一節抽離
    for r in range(2, ws.max_row + 1):
        grp, code, s, t = (ws.cell(r, c).value for c in (1, 2, 3, 4))
        if not (code and s and t):
            continue
        day, p = ws.cell(r, 5).value, ws.cell(r, 6).value
        overlay.append({"grp": str(grp or "列%d" % r), "class": str(code), "subj": s,
                        "t": str(t), "day": str(day) if day else None,
                        "p": int(p) if p else None})
    d["overlay"] = overlay

    # ── 本土語分組 → 自動推導場地封鎖與協同教師限制（與手動填寫取聯集）──
    notes = []
    grade_of = {c["code"]: c["grade"] for c in classes}
    band = {}  # 年級 -> (day, p)
    for lk in locks:
        if lk["subj"] == "本土語文" and lk["class"] in grade_of:
            band[grade_of[lk["class"]]] = (lk["day"], lk["p"])
    live_room = next((rid for rid, nm in d["room_names"].items()
                      if "電腦" in nm or "資訊" in nm), None)
    pool = sorted(d["room_prio"], key=d["room_prio"].get)  # 徵用順位
    ws = wb["本土語分組"]
    for r in range(2, ws.max_row + 1):
        g, lang = ws.cell(r, 1).value, ws.cell(r, 4).value
        if not g or not lang:
            continue
        p_raw = ws.cell(r, 3).value
        if str(p_raw) == "早自習":
            continue  # H16：早自習不佔正課
        try:
            g = int(g)
        except (TypeError, ValueError):
            continue
        if g not in band:
            notes.append(f"本土語分組：{g}年級無固定課鎖定時段，該列未推導（{lang}）")
            continue
        day, p = band[g]
        room = ws.cell(r, 7).value
        if "直播" in str(lang):
            room = room or live_room
        if not room:  # 實體分組未指定 → 依徵用順位自動選
            for rid in pool:
                if (rid, day, p) not in blocked:
                    room = rid
                    notes.append(f"本土語分組：{g}年級{lang} 自動徵用 {d['room_names'].get(rid, rid)}（週{day}第{p}節）")
                    break
        if room and room != "R00":
            blocked.add((room, str(day), int(p)))
        co = ws.cell(r, 9).value
        if co:
            tlimit.add((str(co), str(day), int(p)))
    d["room_blocked"], d["teacher_limit"] = blocked, tlimit
    d["derived_notes"] = notes
    return d


def solve(d, time_limit=60, auto_schedule_tutor=False):
    m = cp_model.CpModel()
    classes, subjects, assign = d["classes"], d["subjects"], d["assign"]
    grade_slot = d["grade_slot"]
    warn = list(d.get("derived_notes", []))

    locked_set = {(lk["class"], lk["subj"]) for lk in d["locks"]}
    lock_teachers = {(lk["class"], lk["subj"]): _text(lk.get("teacher"))
                     for lk in d["locks"] if _text(lk.get("teacher"))}
    native_enabled = d.get("native_lock_enabled")
    if native_enabled is None:
        native_enabled = any(lock["subj"] == "本土語文" for lock in d["locks"])
    overlay_set = {
        (code, subject)
        for ov in d["overlay"]
        for code in _resource_sources(ov)
        for subject in _resource_pull_subjects(ov)
    }
    minnan_group_sources = _minnan_group_sources(d.get("native_groups"))
    tasks = {}
    pool = defaultdict(list)  # 導師自排科目池：class -> [(subj, hours, teacher)]
    missing_courses = []
    required_total = 0
    for c in classes:
        code, g = c["code"], c["grade"]
        for s, info in subjects.items():
            h = info["hours"][g]
            if h == 0:
                continue
            required_total += h
            native_lock = bool(native_enabled) and s == "本土語文" and (code, s) in locked_set
            native_group_owned = native_lock and code in minnan_group_sources
            t = "" if native_group_owned else (lock_teachers.get((code, s)) or assign.get((code, s)))
            mode = d.get("assignment_modes", {}).get((code, s))
            self_arrange = mode == "tutor" if mode else info["self_arrange"]
            if self_arrange and (code, s) not in locked_set:
                # 引擎排課例外：資源班抽離科目、授課者非導師本人。
                bind = ((code, s) in overlay_set or (t and t != c["tutor"]))
                if auto_schedule_tutor:
                    t = t or c.get("tutor") or ""
                elif not bind:
                    pool[code].append((s, h, t or ""))
                    continue
                if bind:
                    reason = ("資源班綁課" if (code, s) in overlay_set
                              else f"授課者{t}非導師")
                    warn.append(f"引擎接手導師科目：{code} {s}（{reason}）")
            if not t and (code, s) not in locked_set:
                warn.append(f"未配教師，未排：{code} {s} {h}節（H12）")
                missing_courses.append({"class": code, "subject": s, "hours": h,
                                        "reason": "未配教師"})
                continue
            if not t:
                warn.append(f"鎖定課未配教師，仍依鎖定排入：{code} {s}（H10>H12，請補配）")
            room = "R00" if native_lock else d["room_override"].get((code, s), info["room"])
            tasks[(code, s)] = {"h": h, "t": t, "room": room, "grade": g, "info": info}
    d["pool"] = pool
    d["missing_courses"] = missing_courses
    d["required_total"] = required_total
    d["auto_schedule_tutor"] = bool(auto_schedule_tutor)

    slots = [(day, p) for day in DAYS for p in PERIODS]
    x = {}
    for (code, s), tk in tasks.items():
        for day, p in slots:
            ok = grade_slot.get((tk["grade"], day, p), False) and p not in tk["info"]["banned"]
            if ok and (tk["grade"], day, p) in d.get("grade_limit", set()):
                ok = False  # H17 學年共同時間
            if ok and (code, day, p) in d.get("class_limit", set()):
                ok = False  # H17 班級限制
            if ok and tk["t"] and (tk["t"], day, p) in d["teacher_limit"]:
                ok = False  # H07
            if ok and tk["room"] != "R00" and (tk["room"], day, p) in d["room_blocked"]:
                ok = False  # H15
            x[(code, s, day, p)] = m.NewBoolVar(f"x_{code}_{s}_{day}{p}") if ok else None

    def var(code, s, day, p):
        return x.get((code, s, day, p))

    def active(code, s):
        return [(day, p, x[(code, s, day, p)]) for day, p in slots if x.get((code, s, day, p)) is not None]

    for (code, s), tk in tasks.items():
        vs = [v for _, _, v in active(code, s)]
        if len(vs) < tk["h"]:
            primary = _diagnostic(
                f"{code} {s} 可用時段不足",
                f"需要 {tk['h']} 節，但依年段、班級、教師、科目及場地限制只剩 {len(vs)} 節可用。",
                "前往不排課時間放寬限制，或調整這門課的授課教師與場地。", "lim")
            diagnostics = [primary] + diagnose_infeasibility(d, tasks, x)
            raise InfeasibleScheduleError(
                f"{code} {s} 可用時段不足（需 {tk['h']} 節，僅剩 {len(vs)} 節）",
                diagnostics[:10])
        m.Add(sum(vs) == tk["h"])

    # H02 班級不衝堂
    for c in classes:
        code = c["code"]
        for day, p in slots:
            vs = [v for (cc, s, dd, pp), v in x.items() if v is not None and cc == code and dd == day and pp == p]
            if vs:
                m.Add(sum(vs) <= 1)

    # H01 教師不衝堂
    teacher_slot = defaultdict(list)
    for (code, s, day, p), v in x.items():
        if v is not None and tasks[(code, s)]["t"]:
            teacher_slot[(tasks[(code, s)]["t"], day, p)].append(v)
    for vs in teacher_slot.values():
        if len(vs) > 1:
            m.Add(sum(vs) <= 1)

    # H03 場地容量（先蒐集；2+1「連堂科任+單節原班」之科目改以連堂變數計占用，約束於連堂建模後生效）
    room_slot = defaultdict(list)
    for (code, s, day, p), v in x.items():
        tk = tasks[(code, s)]
        if v is None or tk["room"] == "R00":
            continue
        if tk["info"]["block"] == "2+1" and tk["info"]["pair_mode"] == "連堂科任+單節原班":
            continue  # 由連堂變數計入
        room_slot[(tk["room"], day, p)].append(v)

    # H10/H14 固定課鎖定
    invalid_locks = []
    for lk in d["locks"]:
        v = var(lk["class"], lk["subj"], lk["day"], lk["p"])
        if v is None:
            invalid_locks.append(lk)
            warn.append(f"鎖定課違反其他硬規則，無法排入：{lk['class']} {lk['subj']} 週{lk['day']}第{lk['p']}節")
        else:
            m.Add(v == 1)
    if invalid_locks:
        raise InfeasibleScheduleError(
            "固定課與其他硬規則衝突",
            diagnose_infeasibility(d, tasks, x, invalid_locks))

    # H13 資源班 overlay
    ov_z = {}
    ov_by_teacher = defaultdict(list)
    for i, ov in enumerate(d["overlay"]):
        sources = _resource_sources(ov)
        pull_subjects = _resource_pull_subjects(ov)
        cand = []
        for day in DAYS:
            for p in PERIODS:
                source_vars = {}
                for code in sources:
                    values = [var(code, subject, day, p) for subject in pull_subjects]
                    source_vars[code] = [value for value in values if value is not None]
                if any(not values for values in source_vars.values()):
                    continue
                if ov["day"] and day != ov["day"]:
                    continue
                if ov["p"] and p != ov["p"]:
                    continue
                if (ov["t"], day, p) in d["teacher_limit"]:
                    continue
                z = m.NewBoolVar(f"ov{i}_{day}{p}")
                for values in source_vars.values():
                    m.Add(z <= sum(values))
                cand.append((day, p, z))
        if not cand:
            source_label = "、".join(sources)
            raise InfeasibleScheduleError(
                f"資源班分組沒有共同可抽離時段：{ov['grp']}（{source_label}）",
                [{"rule": "H13", "message": "請檢查原班可抽離科目、固定時段及資源班教師限制"}])
        m.Add(sum(z for _, _, z in cand) == 1)
        for day, p, z in cand:
            ov_z[(i, day, p)] = z
            ov_by_teacher[(ov["t"], day, p)].append(z)
    for (t, day, p), zs in ov_by_teacher.items():
        m.Add(sum(zs) + sum(teacher_slot.get((t, day, p), [])) <= 1)
    grp_rows = defaultdict(list)
    for i, ov in enumerate(d["overlay"]):
        group_id = str(ov.get("id") or "").rsplit("-", 1)[0] or ov["grp"]
        grp_rows[group_id].append(i)
    for ids in grp_rows.values():
        if len(ids) > 1:
            slotmap = defaultdict(list)
            for (ii, day, p), z in ov_z.items():
                if ii in ids:
                    slotmap[(day, p)].append(z)
            for zs in slotmap.values():
                if len(zs) > 1:
                    m.Add(sum(zs) <= 1)

    # 連堂與分散（H08 自然2+1、H09 視藝2連堂）
    for (code, s), tk in tasks.items():
        info, h = tk["info"], tk["h"]
        acts = active(code, s)
        day_has = {}
        for day in DAYS:
            vs = [v for dd, _, v in acts if dd == day]
            if vs:
                b = m.NewBoolVar(f"d_{code}_{s}_{day}")
                m.AddMaxEquality(b, vs)
                day_has[day] = (b, vs)

        if info["block"] == "2+1" and h == 3:
            pairs = []
            for day in DAYS:
                for p1, p2 in PAIR_START:
                    v1, v2 = var(code, s, day, p1), var(code, s, day, p2)
                    if v1 is not None and v2 is not None:
                        pb = m.NewBoolVar(f"pr_{code}_{s}_{day}{p1}")
                        m.Add(v1 == 1).OnlyEnforceIf(pb)
                        m.Add(v2 == 1).OnlyEnforceIf(pb)
                        pairs.append((day, pb))
                        if info["pair_mode"] == "連堂科任+單節原班" and tk["room"] != "R00":
                            room_slot[(tk["room"], day, p1)].append(pb)
                            room_slot[(tk["room"], day, p2)].append(pb)
            m.Add(sum(pb for _, pb in pairs) == 1)
            for day, (b, vs) in day_has.items():
                day_pairs = [pb for dd, pb in pairs if dd == day]
                two = m.NewBoolVar(f"two_{code}_{s}_{day}")
                if day_pairs:
                    m.AddMaxEquality(two, day_pairs)
                else:
                    m.Add(two == 0)
                m.Add(sum(vs) == 2).OnlyEnforceIf(two)
                m.Add(sum(vs) <= 1).OnlyEnforceIf(two.Not())
            m.Add(sum(b for b, _ in day_has.values()) == 2)
        elif info["block"].startswith("2連堂") and h == 2:
            pairs = []
            for day in DAYS:
                for p1, p2 in PAIR_START:
                    v1, v2 = var(code, s, day, p1), var(code, s, day, p2)
                    if v1 is not None and v2 is not None:
                        pb = m.NewBoolVar(f"pr_{code}_{s}_{day}{p1}")
                        m.Add(v1 == 1).OnlyEnforceIf(pb)
                        m.Add(v2 == 1).OnlyEnforceIf(pb)
                        pairs.append(pb)
            m.Add(sum(pairs) == 1)
        else:
            per_day = 2 if h > 5 else 1
            for day, (b, vs) in day_has.items():
                m.Add(sum(vs) <= per_day)
            if info["spread"]:
                m.Add(sum(b for b, _ in day_has.values()) >= min(int(info["spread"]), h))

    # H03 場地容量約束（含 2+1 連堂占用）
    for (rid, day, p), vs in room_slot.items():
        m.Add(sum(vs) <= d["rooms"].get(rid, 1))

    # ---------------- 軟規則 ----------------
    pen = []
    R = d["rules"]

    def w(rid, default=0):
        return R.get(rid, {}).get("w", default) if R.get(rid, {}).get("on") else 0

    for (code, s), tk in tasks.items():
        rid = "S01" if s == "國語文" else "S02" if s == "數學" else None
        if rid and w(rid):
            for day, p, v in active(code, s):
                if p not in MORNING:
                    pen.append(w(rid) * v)

    for (code, s), tk in tasks.items():
        rid = "S03" if s == "體育" else "S04" if s == "自然科學" else None
        if rid and w(rid):
            db = {}
            for day in DAYS:
                vs = [v for dd, _, v in active(code, s) if dd == day]
                if vs:
                    b = m.NewBoolVar(f"sd_{code}_{s}_{day}")
                    m.AddMaxEquality(b, vs)
                    db[day] = b
            for d1, d2 in zip(DAYS, DAYS[1:]):
                if d1 in db and d2 in db:
                    both = m.NewBoolVar(f"cc_{code}_{s}_{d1}")
                    m.AddMinEquality(both, [db[d1], db[d2]])
                    pen.append(w(rid) * both)

    # S05 行政教師空堂集中
    if w("S05"):
        admin_half = defaultdict(list)
        admin_load = defaultdict(int)
        for (code, s, day, p), v in x.items():
            if v is None:
                continue
            t = tasks[(code, s)]["t"]
            if t and d["roster"].get(t) in ("組長", "主任"):
                admin_half[(t, day, "AM" if p <= 4 else "PM")].append(v)
        for (code, s), tk in tasks.items():
            if tk["t"] and d["roster"].get(tk["t"]) in ("組長", "主任"):
                admin_load[tk["t"]] += tk["h"]
        blocks = defaultdict(list)
        for (t, day, half), vs in admin_half.items():
            b = m.NewBoolVar(f"blk_{t}_{day}{half}")
            m.AddMaxEquality(b, vs)
            blocks[t].append(b)
        for t, bs in blocks.items():
            minb = -(-admin_load[t] // 4)
            ex = m.NewIntVar(0, 10, f"exb_{t}")
            m.Add(ex >= sum(bs) - minb)
            pen.append(w("S05") * ex)

    # S07 減少相鄰節次跨場地移動（聚合式：教師×時段×場地型別；R00 間不計）
    if w("S07"):
        t_rooms = defaultdict(set)
        for (code, s), tk in tasks.items():
            if tk["t"]:
                t_rooms[tk["t"]].add(tk["room"])
        u = {}
        for (code, s, day, p), v in x.items():
            if v is None:
                continue
            t = tasks[(code, s)]["t"]
            if t and len(t_rooms[t]) > 1:
                u.setdefault((t, day, p, tasks[(code, s)]["room"]), []).append(v)
        uvar = {}
        for key, vs in u.items():
            b = m.NewBoolVar("u_%s_%s%s_%s" % key)
            m.AddMaxEquality(b, vs)
            uvar[key] = b
        for t, rset in t_rooms.items():
            if len(rset) <= 1:
                continue
            for day in DAYS:
                for p1, p2 in PAIR_START:
                    for r1 in rset:
                        for r2 in rset:
                            if r1 == r2:
                                continue
                            b1, b2 = uvar.get((t, day, p1, r1)), uvar.get((t, day, p2, r2))
                            if b1 is not None and b2 is not None:
                                mv = m.NewBoolVar(f"mv_{t}_{day}{p1}_{r1}{r2}")
                                m.Add(mv >= b1 + b2 - 1)
                                pen.append(w("S07") * mv)

    # S08/S09 每日負荷；固定的本土語授課與協同節數也納入計算。
    soft_caps = {"科任": 5, "導師": 4, "組長": 4, "主任": 4,
                 "導師兼組長": 4, "導師兼主任": 4,
                 "資源班教師": 5, "資源班導師": 4,
                 "專任輔導教師": 6, "教支人員": 6,
                 "鐘點教師": 6, "其他": 6, "": 6}
    policy_hard_cap = schedule_policy.daily_hard_cap(d)
    tload = defaultdict(list)
    for (code, s, day, p), v in x.items():
        if v is not None and tasks[(code, s)]["t"]:
            tload[(tasks[(code, s)]["t"], day)].append(v)
    native_daily = defaultdict(int)
    for group in d.get("native_groups", []):
        for teacher in (group.get("t"), group.get("assistant")):
            if teacher:
                native_daily[(teacher, group.get("d"))] += 1
    for t, day in set(tload) | set(native_daily):
        vs, fixed = tload[(t, day)], native_daily[(t, day)]
        soft = min(soft_caps.get(d["roster"].get(t, ""), policy_hard_cap), policy_hard_cap)
        m.Add(sum(vs) + fixed <= policy_hard_cap)
        rid = "S09" if d["roster"].get(t) == "導師" else "S08"
        if w(rid) and soft < 7 and len(vs) + fixed > soft:
            over = m.NewIntVar(0, 7, f"ov_{t}_{day}")
            m.Add(over >= sum(vs) + fixed - soft)
            pen.append(w(rid) * over)

    # S06 偏好（保留去年課表）
    for pf in d["prefs"]:
        v = var(pf["class"], pf["subj"], pf["day"], pf["p"])
        if v is not None and w("S06"):
            pen.append(w("S06") * (1 - v))

    m.Minimize(sum(pen))
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit
    solver.parameters.num_workers = min(8, max(1, int(os.getenv("SCHEDULE_SOLVER_WORKERS", "2"))))
    solver.parameters.random_seed = int(os.getenv("SCHEDULE_RANDOM_SEED", "42"))
    status = solver.Solve(m)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        status_name = solver.StatusName(status)
        message = ("求解時間內尚未找到可行解" if status_name == "UNKNOWN"
                   else "無可行解，硬規則彼此衝突")
        raise InfeasibleScheduleError(
            message, diagnose_infeasibility(d, tasks, x, status=status_name), status_name)

    sched = {}
    for (code, s, day, p), v in x.items():
        if v is not None and solver.Value(v):
            tk = tasks[(code, s)]
            sched[(code, day, p)] = (s, tk["t"] or "", tk["room"])
    # 2+1「連堂科任+單節原班」：單節那天場地改標原班 R00
    for (code, s), tk in tasks.items():
        if tk["info"]["block"] == "2+1" and tk["info"]["pair_mode"] == "連堂科任+單節原班":
            by_day = defaultdict(list)
            for (cc, dd, pp), (ss, tt, _) in list(sched.items()):
                if cc == code and ss == s:
                    by_day[dd].append(pp)
            for dd, ps in by_day.items():
                if len(ps) == 1:
                    ss, tt, _ = sched[(code, dd, ps[0])]
                    sched[(code, dd, ps[0])] = (ss, tt, "R00")
    ov_sched = []
    for (i, day, p), z in ov_z.items():
        if solver.Value(z):
            ov = d["overlay"][i]
            for code in _resource_sources(ov):
                pull_subject = next((subject for subject in _resource_pull_subjects(ov)
                                     if var(code, subject, day, p) is not None
                                     and solver.Value(var(code, subject, day, p))), "")
                ov_sched.append((ov.get("id") or f"resource-{i + 1}", ov["grp"], code,
                                 ov["subj"], pull_subject, ov["t"], day, p))
    pool_total = sum(hours for rows in pool.values() for _, hours, _ in rows)
    missing_total = sum(item["hours"] for item in missing_courses)
    remaining_total = max(0, required_total - len(sched))
    objective = solver.ObjectiveValue()
    bound = solver.BestObjectiveBound()
    gap = 0.0 if objective == bound else abs(objective - bound) / max(1.0, abs(objective))
    completion = "complete" if remaining_total == 0 else "partial"
    meta = {
        "status": solver.StatusName(status), "penalty": objective,
        "best_bound": bound, "relative_gap": round(gap, 6),
        "wall": round(solver.WallTime(), 1), "conflicts": solver.NumConflicts(),
        "branches": solver.NumBranches(), "required_total": required_total,
        "scheduled_total": len(sched), "remaining_total": remaining_total,
        "pool_total": pool_total, "missing_total": missing_total,
        "missing_course_count": len(missing_courses), "completion": completion,
        "weekly_cap_violations": list(d.get("weekly_cap_violations", [])),
        "policy": d.get("policy_meta") or d.get("policy") or {},
        "compliance_blocking_issues": list(d.get("compliance_blocking_issues", [])),
        "compliance_warnings": list(d.get("compliance_warnings", [])),
        "auto_schedule_tutor": bool(auto_schedule_tutor),
    }
    return sched, tasks, warn, meta, ov_sched


def validate(d, sched, tasks, ov_sched=()):
    errs = []
    grade_of = {c["code"]: c["grade"] for c in d["classes"]}
    ovt = defaultdict(list)
    for group_id, grp, code, s, pull_subject, t, day, p in ov_sched:
        got = sched.get((code, day, p))
        if not got or got[0] != pull_subject:
            errs.append(f"H13違反(時段不符原班)：{grp} {code} {pull_subject} 週{day}{p}")
        ovt[(t, day, p)].append(group_id)
    for (t, day, p), grps in ovt.items():
        n = len(set(grps)) + sum(1 for (cc, dd, pp), (ss, tt, _) in sched.items()
                                 if tt == t and dd == day and pp == p)
        if n > 1:
            errs.append(f"H13違反(資源班教師衝堂)：{t} 週{day}{p} {grps}")
    tslot, cslot, rslot, tday = defaultdict(list), defaultdict(list), defaultdict(list), defaultdict(list)
    for (code, day, p), (s, t, room) in sched.items():
        g = grade_of[code]
        cslot[(code, day, p)].append(s)
        if t:
            tslot[(t, day, p)].append(f"{code}{s}")
            tday[(t, day)].append(f"{code}{s}")
        if room != "R00":
            rslot[(room, day, p)].append(code)
        if not d["grade_slot"][(g, day, p)]:
            errs.append(f"H05違反：{code} {s} 週{day}{p}")
        if p in tasks[(code, s)]["info"]["banned"]:
            errs.append(f"H06違反：{code} {s} 週{day}{p}")
        if t and (t, day, p) in d["teacher_limit"]:
            errs.append(f"H07違反：{t} 週{day}{p}")
        if (g, day, p) in d.get("grade_limit", set()):
            errs.append(f"H17違反(學年共同時間)：{code} 週{day}{p}")
        if (code, day, p) in d.get("class_limit", set()):
            errs.append(f"H17違反(班級限制)：{code} 週{day}{p}")
        if room != "R00" and (room, day, p) in d["room_blocked"]:
            errs.append(f"H15違反：{room} 週{day}{p}")
    for k, v in tslot.items():
        if len(v) > 1:
            errs.append(f"H01違反：{k} {v}")
    resource_daily_seen = set()
    for group_id, grp, code, s, pull_subject, t, day, p in ov_sched:
        marker = (group_id, t, day, p)
        if t and marker not in resource_daily_seen:
            resource_daily_seen.add(marker)
            tday[(t, day)].append(f"{s}({grp})")
    for group in d.get("native_groups", []):
        for teacher in (group.get("t"), group.get("assistant")):
            if teacher:
                tday[(teacher, group.get("d"))].append(f"{group.get('grp', '本土語分組')}(固定)")
    daily_caps = {role: schedule_policy.daily_hard_cap(d) for role in (
        "科任", "導師", "組長", "主任", "導師兼組長", "導師兼主任",
        "資源班教師", "資源班導師", "專任輔導教師", "教支人員",
        "鐘點教師", "其他", "")}
    for (teacher, day), items in tday.items():
        cap = daily_caps.get(d["roster"].get(teacher, ""), 7)
        if len(items) > cap:
            errs.append(f"每日上限違反：{teacher} 週{day} {len(items)}/{cap}節 {items}")
    for k, v in cslot.items():
        if len(v) > 1:
            errs.append(f"H02違反：{k} {v}")
    for (room, day, p), v in rslot.items():
        if len(v) > d["rooms"].get(room, 1):
            errs.append(f"H03違反：{room} 週{day}{p} {v}")
    for (code, s), tk in tasks.items():
        n = sum(1 for (cc, dd, pp), (ss, _, _) in sched.items() if cc == code and ss == s)
        if n != tk["h"]:
            errs.append(f"節數不符：{code} {s} 排{n}/需{tk['h']}")
    for lk in d["locks"]:
        got = sched.get((lk["class"], lk["day"], lk["p"]))
        if not got or got[0] != lk["subj"]:
            errs.append(f"H10違反：{lk['class']} {lk['subj']} 週{lk['day']}{lk['p']}")
    for (code, s), tk in tasks.items():
        if tk["info"]["block"] != "2+1":
            continue
        cells = sorted((dd, pp) for (cc, dd, pp), (ss, _, _) in sched.items() if cc == code and ss == s)
        days = {dd for dd, _ in cells}
        if len(days) != 2:
            errs.append(f"H08違反(未分兩天)：{code} {s} {cells}")
        for day in days:
            ps = sorted(pp for dd, pp in cells if dd == day)
            if len(ps) == 2 and (ps[1] - ps[0] != 1 or (ps[0] <= 4 < ps[1])):
                errs.append(f"H08違反(連堂/跨午休)：{code} {s} 週{day}{ps}")
    for (code, s), tk in tasks.items():
        if not tk["info"]["block"].startswith("2連堂") or tk["h"] != 2:
            continue
        cells = sorted((dd, pp) for (cc, dd, pp), (ss, _, _) in sched.items()
                       if cc == code and ss == s)
        if len(cells) != 2 or cells[0][0] != cells[1][0] or cells[1][1] - cells[0][1] != 1 \
                or cells[0][1] == 4:
            errs.append(f"H09違反(未連堂/跨午休)：{code} {s} {cells}")
    return errs


def write_output(path, d, sched, tasks, warn, meta, errs, ov_sched=()):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    F = "Microsoft JhengHei"
    HDR = Font(name=F, bold=True, color="FFFFFF", size=10)
    FILL = PatternFill("solid", start_color="305496")
    THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "摘要"
    completion_label = "完整" if meta.get("completion") == "complete" else "部分完成"
    resource_session_count = len({(group_id, teacher, day, period)
                                  for group_id, _, _, _, _, teacher, day, period in ov_sched})
    lines = [f"求解狀態：{meta['status']}　完整度：{completion_label}　penalty={meta['penalty']}　耗時{meta['wall']}秒",
             f"需求：{meta.get('required_total', len(sched))}節　已排：{len(sched)}節　待完成：{meta.get('remaining_total', 0)}節　資源班：{resource_session_count}節",
             f"未配教師：{meta.get('missing_total', 0)}節　導師自排池：{meta.get('pool_total', 0)}節　週上限問題：{len(meta.get('weekly_cap_violations', []))}項",
             f"獨立硬規則檢核：{'零違反 ✓' if not errs else f'{len(errs)} 項違反 ✗'}　最佳界：{meta.get('best_bound', 0)}　gap={meta.get('relative_gap', 0)}"]
    for i, t in enumerate(lines, 1):
        ws.cell(i, 1, excel_safe(t)).font = Font(name=F, size=11, bold=True)
    ws.column_dimensions["A"].width = 80

    def grid(ws, r0, title, cellmap):
        ws.cell(r0, 1, excel_safe(title)).font = Font(name=F, size=11, bold=True)
        for j, day in enumerate(DAYS):
            c = ws.cell(r0 + 1, 2 + j, day); c.font = HDR; c.fill = FILL; c.alignment = CTR; c.border = THIN
        rr = r0 + 2
        for p in PERIODS:
            if p == 5:
                ws.cell(rr, 1, "午休").font = Font(name=F, size=9, italic=True)
                rr += 1
            c = ws.cell(rr, 1, p); c.font = HDR; c.fill = FILL; c.alignment = CTR; c.border = THIN
            for j, day in enumerate(DAYS):
                v = cellmap.get((day, p), "")
                c = ws.cell(rr, 2 + j, excel_safe(v)); c.font = Font(name=F, size=9); c.alignment = CTR; c.border = THIN
            rr += 1
        return rr + 1

    ov_mark = {(code, day, p) for _, _, code, _, _, _, day, p in ov_sched}
    ws = wb.create_sheet("班級課表")
    for c in range(1, 7):
        ws.column_dimensions[chr(64 + c)].width = 14
    r = 1
    for cl in d["classes"]:
        code = cl["code"]
        cm = {}
        for (cc, day, p), (s, t, _) in sched.items():
            if cc == code:
                mark = "※" if (code, day, p) in ov_mark else ""
                cm[(day, p)] = (f"{s}{mark}\n{t}" if t else s + mark)
        if cm:
            r = grid(ws, r, f"{code} 功課表", cm)

    ws = wb.create_sheet("教師課表")
    for c in range(1, 7):
        ws.column_dimensions[chr(64 + c)].width = 14
    teachers = sorted({t for (_, _, _), (_, t, _) in sched.items() if t})
    r = 1
    for t in teachers:
        cm = {}
        for (code, day, p), (s, tt, _) in sched.items():
            if tt == t:
                cm[(day, p)] = f"{s}\n{code}"
        r = grid(ws, r, f"{t} 課表", cm)

    if ov_sched:
        ws = wb.create_sheet("資源班課表(overlay)")
        for c in range(1, 7):
            ws.column_dimensions[chr(64 + c)].width = 16
        r = 1
        for t in sorted({t for _, _, _, _, _, t, _, _ in ov_sched}):
            cm = {}
            grouped = defaultdict(list)
            for group_id, grp, code, s, pull_subject, tt, day, p in ov_sched:
                if tt == t:
                    grouped[(group_id, grp, s, day, p)].append(code)
            for (group_id, grp, subject, day, period), codes in grouped.items():
                cm[(day, period)] = f"{subject}\n{'、'.join(codes)}·{grp}"
            r = grid(ws, r, f"{t} 資源班課表（※=抽離，原班課表不變）", cm)

    pool = d.get("pool", {})
    if pool:
        ws = wb.create_sheet("導師自排工作單")
        for c in range(1, 7):
            ws.column_dimensions[chr(64 + c)].width = 15
        r = 1
        tutor_of = {c["code"]: c.get("tutor", "") for c in d["classes"]}
        grade_of = {c["code"]: c["grade"] for c in d["classes"]}
        for cl in d["classes"]:
            code = cl["code"]
            if code not in pool:
                continue
            cm = {}
            for (cc, day, p), (s, t, _) in sched.items():
                if cc == code:
                    cm[(day, p)] = f"{s}\n{t}" if t else s
            for day in DAYS:  # 空位標 □、不可排標 —
                for p in PERIODS:
                    if (day, p) not in cm:
                        cm[(day, p)] = "□" if d["grade_slot"][(grade_of[code], day, p)] else "—"
            r = grid(ws, r, f"{code} 導師自排工作單（導師：{tutor_of.get(code, '')}）　□=可排空位　—=不可排", cm)
            items = "　".join(f"{s} {h}節({t})" if t else f"{s} {h}節" for s, h, t in pool[code])
            ws.cell(r - 1, 1, excel_safe("科目池：" + items)).font = Font(name=F, size=11, bold=True)
            r += 2

    ws = wb.create_sheet("警示")
    ws.column_dimensions["A"].width = 90
    for i, wmsg in enumerate(warn + [f"[檢核] {e}" for e in errs], 1):
        ws.cell(i, 1, excel_safe(wmsg)).font = Font(name=F, size=10)
    wb.save(path)


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "內湖國小_配課表範本_v3.xlsx"
    dst = sys.argv[2] if len(sys.argv) > 2 else "課表輸出.xlsx"
    tl = int(sys.argv[3]) if len(sys.argv) > 3 else 60  # 求解秒數上限
    auto_tutor = os.getenv("AUTO_SCHEDULE_TUTOR", "1").lower() not in {"0", "false", "no"}
    d = load_data(src)
    sched, tasks, warn, meta, ov_sched = solve(
        d, time_limit=tl, auto_schedule_tutor=auto_tutor)
    errs = validate(d, sched, tasks, ov_sched)
    write_output(dst, d, sched, tasks, warn, meta, errs, ov_sched)
    print(f"求解：{meta['status']} 完整度={meta['completion']} penalty={meta['penalty']} "
          f"bound={meta['best_bound']} gap={meta['relative_gap']} {meta['wall']}s")
    npool = sum(h for lst in d.get("pool", {}).values() for _, h, _ in lst)
    print(f"需求 {meta['required_total']} 節；已排 {len(sched)} 節；待完成 {meta['remaining_total']} 節；"
          f"導師自排池 {npool} 節；週上限問題 {len(meta['weekly_cap_violations'])} 項；"
          f"overlay {len(ov_sched)} 節；警示 {len(warn)} 項；獨立檢核違反 {len(errs)} 項")
    for e in errs[:10]:
        print(" !", e)
    if errs or meta["completion"] != "complete" or meta["weekly_cap_violations"]:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
