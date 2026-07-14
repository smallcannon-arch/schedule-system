import json
import sys
from types import SimpleNamespace

import engine
import openai_advisor
import pytest
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from support_paths import FIXTURES


PROJECT = FIXTURES


def test_summary_redacts_teacher_and_class_identifiers():
    data = engine.load_data(PROJECT / "排課母版_v5.xlsx")

    summary = openai_advisor.build_anonymized_summary(data, "王導師希望 1甲 國語文排上午")
    serialized = json.dumps(summary, ensure_ascii=False)

    assert "王導師" not in serialized
    assert "1甲" not in serialized
    assert "教師" in serialized
    assert "班級" in serialized


def test_plan_can_only_change_supported_soft_rules():
    data = engine.load_data(PROJECT / "排課母版_v5.xlsx")
    original_s02 = data["rules"]["S02"]["w"]
    plan = openai_advisor.OpenAIPlan(
        summary="提高國語文上午優先度。",
        adjustments=[
            openai_advisor.RuleAdjustment(rule_id="S01", weight=9, reason="符合使用者目標"),
        ],
        advice=[],
    )

    audit = openai_advisor.apply_plan(data, plan)

    assert data["rules"]["S01"]["w"] == 9
    assert data["rules"]["S02"]["w"] == original_s02
    assert audit[0]["rule_id"] == "S01"


def test_hard_rule_or_excessive_weight_is_rejected():
    with pytest.raises(ValidationError):
        openai_advisor.RuleAdjustment(rule_id="H01", weight=5, reason="不可修改硬規則")
    with pytest.raises(ValidationError):
        openai_advisor.RuleAdjustment(rule_id="S01", weight=99, reason="超出上限")


def test_responses_api_call_uses_structured_output_and_no_storage(monkeypatch):
    data = engine.load_data(PROJECT / "排課母版_v5.xlsx")
    expected = openai_advisor.OpenAIPlan(
        summary="平衡核心科目與教師負荷。",
        adjustments=[],
        advice=[],
    )
    captured = {}

    class FakeResponses:
        def parse(self, **kwargs):
            captured.update(kwargs)
            return SimpleNamespace(output_parsed=expected)

    class FakeOpenAI:
        def __init__(self, **kwargs):
            captured["client"] = kwargs
            self.responses = FakeResponses()

    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    monkeypatch.setitem(sys.modules, "openai", SimpleNamespace(OpenAI=FakeOpenAI))

    result = openai_advisor.plan_soft_rules(data, "核心科目排上午", "gpt-5.4-mini")

    assert result is expected
    assert captured["model"] == "gpt-5.4-mini"
    assert captured["text_format"] is openai_advisor.OpenAIPlan
    assert captured["store"] is False
    assert captured["max_output_tokens"] == 1800


def test_openai_sheet_neutralizes_formula_like_model_text(tmp_path):
    path = tmp_path / "result.xlsx"
    Workbook().save(path)
    plan = openai_advisor.OpenAIPlan(
        summary="=HYPERLINK(\"https://example.invalid\")",
        adjustments=[
            openai_advisor.RuleAdjustment(
                rule_id="S01", weight=4, reason="+SUM(A1:A2)"),
        ],
        advice=["@SUM(A1:A2)"],
    )

    openai_advisor.append_plan_sheet(
        path, model="=MODEL()", status="applied", plan=plan,
        audit_rows=[{"rule_id": "S01", "before": 2, "after": 4,
                     "reason": "+SUM(A1:A2)"}],
    )

    sheet = load_workbook(path, data_only=False)["OpenAI規劃"]
    formula_like = [cell for row in sheet.iter_rows() for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("'")]
    assert len(formula_like) >= 4
    assert all(cell.data_type == "s" for cell in formula_like)
