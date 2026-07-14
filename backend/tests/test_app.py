import base64
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import app
import openai_advisor
from support_paths import FIXTURES


PROJECT = FIXTURES
CLIENT = TestClient(app.app)


def test_rejects_non_xlsx_extension():
    response = CLIENT.post("/solve", files={"file": ("bad.txt", b"hello", "text/plain")})

    assert response.status_code == 415
    assert set(response.json()) == {"error"}


def test_rejects_invalid_xlsx_without_exposing_traceback():
    response = CLIENT.post(
        "/solve",
        files={"file": ("bad.xlsx", b"not-an-excel-file", app.XLSX_MIME)},
    )

    assert response.status_code == 422
    assert set(response.json()) == {"error"}
    assert "trace" not in response.text.lower()


def test_rejects_upload_over_size_limit(monkeypatch):
    monkeypatch.setattr(app, "MAX_UPLOAD_BYTES", 4)

    response = CLIENT.post(
        "/solve",
        files={"file": ("large.xlsx", b"12345", app.XLSX_MIME)},
    )

    assert response.status_code == 413


def test_requires_api_key_when_configured(monkeypatch):
    monkeypatch.setattr(app, "API_KEY", "test-secret")

    denied = CLIENT.post(
        "/solve",
        files={"file": ("bad.xlsx", b"not-excel", app.XLSX_MIME)},
    )
    accepted_key = CLIENT.post(
        "/solve",
        headers={"x-api-key": "test-secret"},
        files={"file": ("bad.xlsx", b"not-excel", app.XLSX_MIME)},
    )

    assert denied.status_code == 401
    assert accepted_key.status_code == 422


def test_rate_limiter_rejects_requests_over_instance_limit(monkeypatch):
    monkeypatch.setattr(app, "RATE_LIMIT_PER_MINUTE", 1)
    with app.RATE_LOCK:
        app.SOLVE_REQUESTS.clear()
    try:
        assert app._claim_rate_limit() is True
        assert app._claim_rate_limit() is False
    finally:
        with app.RATE_LOCK:
            app.SOLVE_REQUESTS.clear()


def test_security_middleware_rejects_oversized_body_before_route_parsing(monkeypatch):
    monkeypatch.setattr(app, "MAX_UPLOAD_BYTES", 32)

    response = CLIENT.post(
        "/solve-data", content=b"x" * (1024 * 1024 + 64),
        headers={"content-type": "application/json"},
    )

    assert response.status_code == 413
    assert response.json() == {"error": "請求資料超過大小上限"}


def test_security_headers_are_present_and_production_docs_are_disabled():
    health = CLIENT.get("/health")
    docs = CLIENT.get("/docs")

    assert health.headers["x-content-type-options"] == "nosniff"
    assert health.headers["x-frame-options"] == "DENY"
    assert health.headers["referrer-policy"] == "no-referrer"
    assert docs.status_code == 404


def test_accepts_v5_and_returns_validated_workbook():
    source = PROJECT / "排課母版_v5.xlsx"
    response = CLIENT.post(
        "/solve",
        files={"file": (source.name, source.read_bytes(), app.XLSX_MIME)},
        data={"time_limit": "10", "strict_complete": "false"},
    )

    assert response.status_code == 200
    assert response.content.startswith(b"PK\x03\x04")
    assert response.headers["x-violations"] == "0"
    assert response.headers["x-schedule-completeness"] == "partial"
    assert int(response.headers["x-tutor-pool"]) > 0


def test_strict_mode_rejects_partial_result_without_inventing_weekly_targets():
    source = PROJECT / "排課母版_v5.xlsx"
    response = CLIENT.post(
        "/solve",
        files={"file": (source.name, source.read_bytes(), app.XLSX_MIME)},
        data={"time_limit": "10", "strict_complete": "true"},
    )

    assert response.status_code == 409
    payload = response.json()
    assert payload["error"] == "課表尚未達到正式完成標準"
    assert payload["completion"] == "partial"
    assert payload["weekly_cap_issues"] == []
    assert any("尚未填寫教師每週基準節數" in item for item in payload["compliance_warnings"])


def test_json_response_contains_workbook_and_structured_schedule():
    source = PROJECT / "排課母版_v5.xlsx"
    response = CLIENT.post(
        "/solve",
        files={"file": (source.name, source.read_bytes(), app.XLSX_MIME)},
        data={"time_limit": "10", "strict_complete": "false",
              "auto_schedule_tutor": "false", "return_json": "true"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert base64.b64decode(payload["workbook_base64"]).startswith(b"PK\x03\x04")
    assert payload["schedule"]
    assert payload["meta"]["auto_schedule_tutor"] is False
    assert payload["meta"]["pool_total"] > 0
    assert all({"code", "day", "period", "subject", "teacher", "room"} <= set(row)
               for row in payload["schedule"])


def test_openai_request_requires_server_configuration(monkeypatch):
    monkeypatch.setattr(app, "OPENAI_ENABLED", False)

    response = CLIENT.post(
        "/solve",
        files={"file": ("input.xlsx", b"placeholder", app.XLSX_MIME)},
        data={"use_openai": "true", "ai_goal": "國數優先排上午"},
    )

    assert response.status_code == 503
    assert "OPENAI_API_KEY" in response.json()["error"]


def test_openai_plan_is_applied_and_written_to_workbook(monkeypatch):
    source = PROJECT / "排課母版_v5.xlsx"
    plan = openai_advisor.OpenAIPlan(
        summary="以上午核心學科與行政空堂集中為優先。",
        adjustments=[
            openai_advisor.RuleAdjustment(rule_id="S01", weight=8, reason="國語文上午優先"),
            openai_advisor.RuleAdjustment(rule_id="S05", weight=5, reason="集中行政空堂"),
        ],
        advice=["先檢查行政教師週節數是否合理。"],
    )
    monkeypatch.setattr(app, "OPENAI_ENABLED", True)
    monkeypatch.setattr(openai_advisor, "plan_soft_rules", lambda data, goal, model: plan)

    response = CLIENT.post(
        "/solve",
        files={"file": (source.name, source.read_bytes(), app.XLSX_MIME)},
        data={"time_limit": "10", "strict_complete": "false", "use_openai": "true", "ai_goal": "國數優先排上午"},
    )

    assert response.status_code == 200
    assert response.headers["x-openai-status"] == "applied"
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert "OpenAI規劃" in workbook.sheetnames
    values = [cell.value for row in workbook["OpenAI規劃"].iter_rows() for cell in row]
    assert "S01" in values
    assert 8 in values


def test_openai_failure_falls_back_to_original_rules(monkeypatch):
    source = PROJECT / "排課母版_v5.xlsx"

    def fail(*args, **kwargs):
        raise RuntimeError("provider unavailable")

    monkeypatch.setattr(app, "OPENAI_ENABLED", True)
    monkeypatch.setattr(openai_advisor, "plan_soft_rules", fail)
    response = CLIENT.post(
        "/solve",
        files={"file": (source.name, source.read_bytes(), app.XLSX_MIME)},
        data={"time_limit": "10", "strict_complete": "false", "use_openai": "true", "ai_goal": "平衡教師負荷"},
    )

    assert response.status_code == 200
    assert response.headers["x-openai-status"] == "failed"
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook["OpenAI規劃"]["B1"].value == "failed"
