from copy import deepcopy

import engine
import pytest
from openpyxl import load_workbook
from support_paths import FIXTURES


V5_TEMPLATE = FIXTURES / "排課母版_v5.xlsx"
V6_TEMPLATE = FIXTURES / "排課母版_v6.xlsx"
V4_TEMPLATE = FIXTURES / "排課系統_通用母版_v4.xlsx"


def test_loads_v5_schema_and_expands_teacher_centered_assignments():
    data = engine.load_data(V5_TEMPLATE)

    assert data["schema_version"] == 5
    assert len(data["classes"]) == 3
    assert data["assign"][("1甲", "國語文")] == "王導師"
    assert data["assign"][("1甲", "音樂")] == "張科任"
    assert ("英語教室", "二", 1) in data["room_blocked"]
    assert ("閩語教支", "二", 1) in data["teacher_limit"]


def test_loads_v6_merged_teacher_and_assignment_sheet():
    data = engine.load_data(V6_TEMPLATE)

    assert data["schema_version"] == 6
    assert "張科任" in data["roster"]
    assert data["assign"][("1甲", "音樂")] == "張科任"
    assert data["assign"][("1甲", "國語文")] == "王導師"


def test_v6_excel_requires_explicit_distributed_native_language_mode(tmp_path):
    workbook = load_workbook(V6_TEMPLATE)
    sheet = workbook["本土語分組"]
    for row in range(2, 5):
        sheet.cell(row, 12, "各語別分開上")
        sheet.cell(row, 13, "國語文、數學")
    values = [
        1, "三", 2, "客語-海陸", "1年級客語海陸組", "1甲", 1,
        "客語教支", None, "原班教室", "實體", "各語別分開上", "國語文、數學",
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(5, column, value)
    path = tmp_path / "distributed-native.xlsx"
    workbook.save(path)

    data = engine.load_data(path)

    assert data["native_arrangement"] == "distributed"
    assert data["native_bands"] == []
    assert data["native_external_sources"] == set()
    assert len(data["native_groups"]) == 4
    assert not any(lock["subj"] == "本土語文" for lock in data["locks"])


def test_v6_excel_does_not_silently_switch_multi_slot_native_language_mode(tmp_path):
    workbook = load_workbook(V6_TEMPLATE)
    sheet = workbook["本土語分組"]
    values = [
        1, "三", 2, "客語-海陸", "1年級客語海陸組", "1甲", 1,
        "客語教支", None, "原班教室", "實體", None, "國語文、數學",
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(5, column, value)
    path = tmp_path / "unmarked-multi-slot-native.xlsx"
    workbook.save(path)

    with pytest.raises(ValueError, match="明確選擇.*各語別分開上"):
        engine.load_data(path)


def test_v6_duplicate_teacher_rows_merge_only_when_metadata_matches(tmp_path):
    workbook = load_workbook(V6_TEMPLATE)
    sheet = workbook["教師與配課"]
    source = [sheet.cell(3, column).value for column in range(1, 8)]
    row = sheet.max_row + 1
    for column, value in enumerate(source, start=1):
        sheet.cell(row, column, value)
    target = tmp_path / "duplicate-teacher-same.xlsx"
    workbook.save(target)

    data = engine.load_data(target)

    assert any("重複，已合併為同一位教師" in note for note in data["derived_notes"])


def test_v6_conflicting_duplicate_teacher_row_is_rejected(tmp_path):
    workbook = load_workbook(V6_TEMPLATE)
    sheet = workbook["教師與配課"]
    row = sheet.max_row + 1
    sheet.cell(row, 1, sheet["A3"].value)
    sheet.cell(row, 2, "組長")
    sheet.cell(row, 3, 20)
    sheet.cell(row, 4, 0)
    target = tmp_path / "duplicate-teacher-conflict.xlsx"
    workbook.save(target)

    with pytest.raises(ValueError, match="重複.*不一致"):
        engine.load_data(target)


@pytest.mark.parametrize(("sheet_name", "values", "message"), [
    ("班級", ["7甲", "七年級", "王導師"], "年級必須是 1 到 6"),
    ("班級", [None, 1, "王導師"], "缺少班級代碼"),
    ("班級", ["1甲", 1, "王導師"], "班級代碼重複"),
    ("場地", [None, 2], "缺少場地名稱"),
    ("場地", ["自然教室", 2], "場地名稱重複"),
    ("場地", ["創客教室", 1.5], "容量必須是正整數"),
    ("科目節數", [None, 1, 0, 0, 0, 0, 0], "缺少科目名稱"),
    ("科目節數", ["閱讀", 1.5, 0, 0, 0, 0, 0], "節數必須是整數"),
    ("年段時段", ["七年級", 1, 1], "年級必須是 1 到 6"),
    ("本土語分組", [1, "六", 1, "客語", "客語教師"], "時段無效"),
])
def test_v6_invalid_nonblank_rows_are_rejected(tmp_path, sheet_name, values, message):
    workbook = load_workbook(V6_TEMPLATE)
    sheet = workbook[sheet_name]
    row = sheet.max_row + 1
    for column, value in enumerate(values, start=1):
        sheet.cell(row, column, value)
    target = tmp_path / f"invalid-{sheet_name}.xlsx"
    workbook.save(target)

    with pytest.raises(ValueError, match=message):
        engine.load_data(target)


@pytest.mark.parametrize(("sheet_name", "values", "message"), [
    ("不排課時間", ["王導師", "六", 1, "不可排"], "星期不正確"),
    ("資源班overlay", ["測試組", "1甲", "國語文", "名冊外教師"], "教師不在名冊"),
])
def test_v6_invalid_optional_rows_are_rejected(tmp_path, sheet_name, values, message):
    workbook = load_workbook(V6_TEMPLATE)
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    if sheet.max_row == 1 and sheet["A1"].value is None:
        headers = (["對象", "星期", "節次", "類型", "備註"] if sheet_name == "不排課時間"
                   else ["組別", "原班", "科目", "資源班教師", "星期", "節次"])
        for column, value in enumerate(headers, start=1):
            sheet.cell(1, column, value)
    sheet.append(values)
    target = tmp_path / f"invalid-{sheet_name}.xlsx"
    workbook.save(target)

    with pytest.raises(ValueError, match=message):
        engine.load_data(target)


def test_v6_partial_assignment_row_is_rejected(tmp_path):
    workbook = load_workbook(V6_TEMPLATE)
    sheet = workbook["教師與配課"]
    row = sheet.max_row + 1
    sheet.cell(row, 9, sheet["A3"].value)
    sheet.cell(row, 11, workbook["班級"]["A2"].value)
    target = tmp_path / "partial-assignment.xlsx"
    workbook.save(target)

    with pytest.raises(ValueError, match="完整填寫教師、科目與任教班級"):
        engine.load_data(target)


def test_v6_zero_hour_assignment_is_rejected(tmp_path):
    workbook = load_workbook(V6_TEMPLATE)
    sheet = workbook["教師與配課"]
    row = sheet.max_row + 1
    sheet.cell(row, 9, sheet["A3"].value)
    sheet.cell(row, 10, "英語文")
    sheet.cell(row, 11, "1甲")
    target = tmp_path / "zero-hour-assignment.xlsx"
    workbook.save(target)

    with pytest.raises(ValueError, match="1甲的英語文節數為 0"):
        engine.load_data(target)


def test_v6_custom_subject_is_loaded_without_alias_mapping(tmp_path):
    workbook = load_workbook(V6_TEMPLATE)
    subjects = workbook["科目節數"]
    school_subject_row = next(
        row for row in range(2, subjects.max_row + 1)
        if subjects.cell(row, 1).value == "校訂A")
    subjects.cell(school_subject_row, 2, 0)
    row = subjects.max_row + 1
    for column, value in enumerate(
            ["閱讀", 1, 0, 0, 0, 0, 0, "原班教室", "否", ""], start=1):
        subjects.cell(row, column, value)
    assignments = workbook["教師與配課"]
    row = assignments.max_row + 1
    assignments.cell(row, 9, assignments["A3"].value)
    assignments.cell(row, 10, "閱讀")
    assignments.cell(row, 11, "1甲")
    target = tmp_path / "custom-subject.xlsx"
    workbook.save(target)

    data = engine.load_data(target)

    assert data["subjects"]["閱讀"]["hours"][1] == 1
    assert data["assign"][("1甲", "閱讀")] == assignments["A3"].value


def test_v6_direct_parser_reads_alias_limits_and_resource_overlay(tmp_path):
    workbook = load_workbook(V6_TEMPLATE)
    teacher_sheet = workbook["教師與配課"]
    teacher = teacher_sheet["A3"].value
    alias_teacher = teacher_sheet["I6"].value
    code = workbook["班級"]["A2"].value
    teacher_sheet["J6"] = "閩南語"

    limit_sheet = (workbook["不排課時間"] if "不排課時間" in workbook.sheetnames
                   else workbook.create_sheet("不排課時間"))
    if limit_sheet.max_row == 1 and limit_sheet["A1"].value is None:
        for column, value in enumerate(["對象", "星期", "節次", "類型", "備註"], start=1):
            limit_sheet.cell(1, column, value)
    limit_sheet.append([teacher, "五", 7, "不可排", "測試限制"])
    overlay_sheet = (workbook["資源班overlay"] if "資源班overlay" in workbook.sheetnames
                     else workbook.create_sheet("資源班overlay"))
    if overlay_sheet.max_row == 1 and overlay_sheet["A1"].value is None:
        for column, value in enumerate(["組別", "原班", "科目", "資源班教師", "星期", "節次"], start=1):
            overlay_sheet.cell(1, column, value)
    overlay_sheet.append(["測試資源組", code, "國語文", teacher, "一", 1])
    target = tmp_path / "v6-options.xlsx"
    workbook.save(target)

    data = engine.load_data(target)

    assert data["assign"][(code, "本土語文")] == alias_teacher
    assert (teacher, "五", 7) in data["teacher_limit"]
    assert data["overlay"] == [{
        "grp": "測試資源組", "class": code, "subj": "國語文", "t": teacher,
        "day": "一", "p": 1,
    }]


def test_v6_template_uses_consistent_fonts_and_bounded_notes():
    workbook = load_workbook(V6_TEMPLATE, data_only=False)

    assert "教師與配課" in workbook.sheetnames
    assert "教師" not in workbook.sheetnames
    assert "配課" not in workbook.sheetnames
    assert workbook["教師與配課"]["A2"].value == "姓名"
    assert workbook["教師與配課"]["E2"].value == "學校 Google 帳號\n(教支人員可選填)"
    assert workbook["教師與配課"]["F2"].value == "可授本土語別\n(可複選)"
    assert workbook["教師與配課"]["B7"].value == "教支人員"
    assert workbook["教師與配課"]["I2"].value == "教師姓名"
    assert workbook["本土語分組"]["E1"].value == "分組名稱"
    assert workbook["本土語分組"]["H1"].value == "授課教師"
    assert workbook["說明"].max_column == 8
    assert any("只填帳號，不需提供密碼" in str(cell.value or "")
               for row in workbook["說明"].iter_rows() for cell in row)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    assert cell.font.name == "Microsoft JhengHei"


def test_loads_frontend_cloud_draft_schema_without_excel():
    slots = [[1, 1, 1, 1, 0, 0, 0] for _ in range(5)]
    payload = {
        "classes": [{"g": 1, "i": 1, "code": "1甲", "tutor": "王老師"}],
        "roster": {"王老師": "導師"},
        "tcap": {"王老師": {"cap": 16, "minus": 0}},
        "rooms": {"R00": 99},
        "subjects": {"國語文": {
            "hours": [1, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
            "block": "", "self": True, "pairMode": "",
        }},
        "gslot": {str(grade): slots for grade in range(1, 7)},
        "assign": {"1甲": {"國語文": "王老師"}},
        "override": {}, "locks": [], "blocked": [], "resGroups": [],
    }

    data = engine.load_frontend_data(
        payload,
        limits=[["王老師", "一", "1", "不可排", "會議"]],
        rules=[["S01", "軟", "偏好", "國語優先上午", "penalty=4", "是"]],
    )
    schedule, _, _, meta, _ = engine.solve(data, time_limit=5, auto_schedule_tutor=True)

    assert data["assign"][("1甲", "國語文")] == "王老師"
    assert ("王老師", "一", 1) in data["teacher_limit"]
    assert data["rules"]["S01"] == {"on": True, "w": 4}
    assert len(schedule) == 1
    assert meta["completion"] == "complete"


def test_soft_rule_quality_report_explains_weighted_penalty():
    closed = [0, 0, 0, 0, 0, 0, 0]
    grade_slots = [[0, 0, 0, 0, 1, 0, 0], closed, closed, closed, closed]
    payload = {
        "classes": [{"g": 1, "i": 1, "code": "1甲", "tutor": "王老師"}],
        "roster": {"王老師": "科任"}, "rooms": {"R00": 99},
        "subjects": {"國語文": {
            "hours": [1, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
            "block": "", "self": False, "pairMode": "",
        }},
        "gslot": {str(grade): grade_slots for grade in range(1, 7)},
        "assign": {"1甲": {"國語文": "王老師"}},
        "override": {}, "locks": [], "blocked": [], "resGroups": [],
    }
    data = engine.load_frontend_data(
        payload, rules=[["S01", "軟", "偏好", "國語優先上午", "penalty=4", "是"]])

    schedule, _, _, meta, _ = engine.solve(data, time_limit=5, auto_schedule_tutor=True)

    assert schedule[("1甲", "一", 5)][0] == "國語文"
    report = {item["rule_id"]: item for item in meta["quality_report"]}
    assert report["S01"]["violations"] == 1
    assert report["S01"]["weighted_penalty"] == 4
    assert report["S01"]["details"] == [{
        "class": "1甲", "subject": "國語文", "day": "一", "period": 5, "units": 1,
    }]
    assert meta["quality_violation_total"] == 1
    assert meta["quality_penalty_total"] == meta["penalty"] == 4


def test_missing_courses_is_canonical_for_tutor_pending_and_missing_teacher():
    grade_slots = [[1, 1, 1, 1, 0, 0, 0] for _ in range(5)]
    payload = {
        "classes": [{"g": 1, "i": 1, "code": "1甲", "tutor": "王老師"}],
        "roster": {"王老師": "導師"}, "rooms": {"R00": 99},
        "subjects": {
            "國語文": {"hours": [1, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
                       "block": "", "self": True, "pairMode": ""},
            "音樂": {"hours": [1, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
                     "block": "", "self": False, "pairMode": ""},
        },
        "gslot": {str(grade): grade_slots for grade in range(1, 7)},
        "assign": {"1甲": {"國語文": "王老師"}},
        "override": {}, "locks": [], "blocked": [], "resGroups": [],
    }
    data = engine.load_frontend_data(payload)

    _, _, _, meta, _ = engine.solve(data, time_limit=5, auto_schedule_tutor=False)

    by_reason = {item["reason_code"]: item for item in meta["missing_courses"]}
    assert by_reason["tutor_pending"] == {
        "class": "1甲", "subject": "國語文", "hours": 1,
        "required": 1, "scheduled": 0,
        "reason_code": "tutor_pending", "reason": "導師待排",
    }
    assert by_reason["missing_teacher"]["subject"] == "音樂"
    assert meta["pool_total"] == 1
    assert meta["missing_total"] == 1
    assert meta["diagnostic_shortfall_total"] == 0
    assert meta["unfinished_course_count"] == 2
    assert meta["remaining_total"] == 2


def _diagnostic_shortfall_payload(block=""):
    closed = [0, 0, 0, 0, 0, 0, 0]
    grade_slots = [[1, 0, 0, 0, 0, 0, 0], closed, closed, closed, closed]
    return {
        "classes": [{"g": 1, "i": 1, "code": "1甲", "tutor": "王老師"}],
        "roster": {"王老師": "科任"}, "rooms": {"R00": 99},
        "subjects": {"國語文": {
            "hours": [2, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
            "block": block, "self": False, "pairMode": "",
        }},
        "gslot": {str(grade): grade_slots for grade in range(1, 7)},
        "assign": {"1甲": {"國語文": "王老師"}},
        "override": {}, "locks": [], "blocked": [], "resGroups": [],
    }


def test_diagnostic_draft_allows_only_ordinary_course_shortfall_and_reports_it(tmp_path):
    data = engine.load_frontend_data(
        _diagnostic_shortfall_payload(), allow_course_shortfall=True)
    data["subjects"]["國語文"]["spread"] = 2

    with pytest.raises(engine.InfeasibleScheduleError, match="可用時段不足"):
        engine.solve(data, time_limit=5)

    schedule, tasks, warnings, meta, overlay = engine.solve(
        data, time_limit=5, diagnostic_draft=True)

    assert len(schedule) == 1
    assert engine.validate(
        data, schedule, tasks, overlay,
        diagnostic_shortfalls={("1甲", "國語文"): 1},
    ) == []
    assert meta["diagnostic_draft"] is True
    assert meta["diagnostic_shortfall_total"] == 1
    assert meta["completion"] == "partial"
    assert meta["missing_courses"] == [{
        "class": "1甲", "subject": "國語文", "hours": 1,
        "required": 2, "scheduled": 1,
        "reason_code": "diagnostic_shortfall", "reason": "診斷草案未排入",
    }]
    assert any("診斷草案未排入" in warning for warning in warnings)

    output = tmp_path / "diagnostic-draft.xlsx"
    engine.write_output(output, data, schedule, tasks, warnings, meta, [], overlay)
    workbook = load_workbook(output)
    assert "診斷草案｜不可發布或作為正式上傳檔" in workbook["摘要"]["A1"].value
    assert workbook["摘要"]["A1"].fill.fgColor.rgb.endswith("F4CCCC")
    unfinished = list(workbook["待完成課程"].values)
    assert any(row[0] == "1甲" and row[1] == "國語文" and row[4] == 1 for row in unfinished[1:])


@pytest.mark.parametrize("block", ["2+1", "2連堂"])
def test_diagnostic_draft_keeps_block_courses_exact(block):
    data = engine.load_frontend_data(
        _diagnostic_shortfall_payload(block), allow_course_shortfall=True)

    with pytest.raises(engine.InfeasibleScheduleError, match="可用時段不足"):
        engine.solve(data, time_limit=5, diagnostic_draft=True)


def test_diagnostic_draft_returns_phase1_solution_when_quality_phase_times_out(monkeypatch):
    data = engine.load_frontend_data(
        _diagnostic_shortfall_payload(), allow_course_shortfall=True)
    real_solver = engine.cp_model.CpSolver
    calls = []

    class TimedOutSolver:
        def __init__(self):
            self.parameters = type("Parameters", (), {})()

        def Solve(self, _model):
            return engine.cp_model.UNKNOWN

        def WallTime(self):
            return 0.0

    def solver_factory():
        calls.append(True)
        return real_solver() if len(calls) == 1 else TimedOutSolver()

    monkeypatch.setattr(engine.cp_model, "CpSolver", solver_factory)

    schedule, _, _, meta, _ = engine.solve(
        data, time_limit=5, diagnostic_draft=True)

    assert len(schedule) == 1
    assert meta["diagnostic_shortfall_total"] == 1
    assert meta["diagnostic_quality_optimized"] is False
    assert meta["diagnostic_fallback_to_phase1"] is True


def _resource_frontend_payload():
    slots = [[1, 1, 1, 1, 0, 0, 0] for _ in range(5)]
    return {
        "classes": [
            {"g": 1, "i": 1, "code": "1甲", "tutor": "王導師"},
            {"g": 1, "i": 2, "code": "1乙", "tutor": "李導師"},
        ],
        "roster": {"王導師": "導師", "李導師": "導師", "資源教師": "資源班教師"},
        "rooms": {"R00": 99},
        "subjects": {
            "國語文": {"hours": [2, 0, 0, 0, 0, 0], "room": "R00",
                       "banned": [], "block": "", "self": True, "pairMode": ""},
            "綜合活動": {"hours": [3, 0, 0, 0, 0, 0], "room": "R00",
                         "banned": [], "block": "", "self": True, "pairMode": ""},
        },
        "gslot": {str(grade): slots for grade in range(1, 7)},
        "assign": {
            "1甲": {"國語文": "王導師", "綜合活動": "王導師"},
            "1乙": {"國語文": "李導師", "綜合活動": "李導師"},
        },
        "override": {}, "locks": [], "blocked": [],
        "resGroups": [{
            "id": "grade-1-a", "grp": "一年級A組", "sources": ["1甲", "1乙"],
            "subj": "國語文", "pullSubjects": ["綜合活動"], "t": "資源教師", "n": 3,
            "scheduleMode": "fixed",
            "slots": [{"d": "一", "p": 1}, {"d": "三", "p": 1}, {"d": "五", "p": 1}],
        }],
    }


@pytest.mark.parametrize(("field", "message"), [
    ("class_grade", "年級必須是 1 到 6"),
    ("room_capacity", "容量必須是正整數"),
    ("subject_hours", "節數必須是整數"),
])
def test_frontend_numeric_fields_reject_fractional_values(field, message):
    payload = _resource_frontend_payload()
    if field == "class_grade":
        payload["classes"][0]["g"] = 1.5
    elif field == "room_capacity":
        payload["rooms"]["資源教室"] = 1.5
    else:
        payload["subjects"]["國語文"]["hours"][0] = 1.5

    with pytest.raises(ValueError, match=message):
        engine.load_frontend_data(payload)


def test_resource_group_combines_classes_and_counts_teacher_sessions_once():
    data = engine.load_frontend_data(_resource_frontend_payload())
    schedule, tasks, _, _, overlay = engine.solve(
        data, time_limit=5, auto_schedule_tutor=False)

    assert len(data["overlay"]) == 3
    assert data["teacher_weekly_load"]["資源教師"] == 3
    assert len(overlay) == 6
    assert len({(row[0], row[6], row[7]) for row in overlay}) == 3
    assert {(row[2], row[4]) for row in overlay} == {
        ("1甲", "綜合活動"), ("1乙", "綜合活動")}
    assert {(code, day, period) for code, day, period in schedule
            if code in {"1甲", "1乙"}} == {
        ("1甲", "一", 1), ("1甲", "三", 1), ("1甲", "五", 1),
        ("1乙", "一", 1), ("1乙", "三", 1), ("1乙", "五", 1),
    }
    assert engine.validate(data, schedule, tasks, overlay) == []


def test_resource_group_uses_the_same_pull_subject_for_every_source_class():
    payload = _resource_frontend_payload()
    payload["resGroups"][0]["pullSubjects"] = ["國語文", "綜合活動"]

    data = engine.load_frontend_data(payload)
    schedule, tasks, _, _, overlay = engine.solve(
        data, time_limit=5, auto_schedule_tutor=False)

    for day in ("一", "三", "五"):
        subjects = {schedule[(code, day, 1)][0] for code in ("1甲", "1乙")}
        assert len(subjects) == 1
        pull_subjects = {row[4] for row in overlay if row[6:] == (day, 1)}
        assert pull_subjects == subjects
    assert engine.validate(data, schedule, tasks, overlay) == []


def test_resource_group_rejects_fixed_slot_with_different_locked_pull_subjects():
    payload = _resource_frontend_payload()
    payload["resGroups"][0].update({
        "pullSubjects": ["國語文", "綜合活動"],
        "n": 1,
        "slots": [{"d": "一", "p": 1}],
    })
    payload["locks"] = [
        {"c": "1甲", "d": "一", "p": 1, "s": "國語文"},
        {"c": "1乙", "d": "一", "p": 1, "s": "綜合活動"},
    ]

    data = engine.load_frontend_data(payload)

    with pytest.raises(engine.InfeasibleScheduleError):
        engine.solve(data, time_limit=5, auto_schedule_tutor=False)


def test_resource_fixed_slots_must_match_weekly_periods():
    payload = _resource_frontend_payload()
    payload["resGroups"][0]["slots"].pop()

    with pytest.raises(ValueError, match="固定時段 2 節，與每週節數 3 不一致"):
        engine.load_frontend_data(payload)


def test_resource_pull_subject_keeps_existing_minnan_subject_name():
    payload = _resource_frontend_payload()
    payload["subjects"]["閩南語"] = payload["subjects"].pop("綜合活動")
    for row in payload["assign"].values():
        row["閩南語"] = row.pop("綜合活動")
    payload["resGroups"][0]["pullSubjects"] = ["閩南語"]

    data = engine.load_frontend_data(payload)

    assert data["overlay"][0]["pull_subjects"] == ["閩南語"]


def test_frontend_lock_prefers_exact_minnan_subject_name():
    payload = _resource_frontend_payload()
    payload["subjects"]["閩南語"] = deepcopy(payload["subjects"]["綜合活動"])
    payload["subjects"]["本土語文"] = deepcopy(payload["subjects"]["國語文"])
    payload["assign"]["1甲"]["閩南語"] = "王導師"
    payload["assign"]["1甲"]["本土語文"] = "王導師"
    payload["locks"] = [{"c": "1甲", "d": "二", "p": 1, "s": "閩南語"}]

    data = engine.load_frontend_data(payload)

    assert data["locks"][0]["subj"] == "閩南語"


def test_resource_fixed_slot_rejects_period_outside_source_grade_schedule():
    payload = _resource_frontend_payload()
    payload["resGroups"][0]["slots"][0] = {"d": "一", "p": 5}

    with pytest.raises(ValueError, match="一年級A組固定時段週一第5節不在1年級可排時段"):
        engine.load_frontend_data(payload)


def test_resource_early_study_is_a_fixed_resource_only_session():
    payload = _resource_frontend_payload()
    payload["resGroups"][0]["slots"] = [
        {"d": "一", "p": 0}, {"d": "三", "p": 1}, {"d": "五", "p": 1},
    ]

    data = engine.load_frontend_data(payload)
    schedule, tasks, _, _, overlay = engine.solve(
        data, time_limit=5, auto_schedule_tutor=False)

    early = [row for row in overlay if row[7] == 0]
    assert len(early) == 2
    assert {row[2] for row in early} == {"1甲", "1乙"}
    assert {row[4] for row in early} == {"早自修"}
    assert engine.validate(data, schedule, tasks, overlay) == []


def test_resource_early_study_rejects_teacher_double_booking():
    payload = _resource_frontend_payload()
    first = payload["resGroups"][0]
    first["n"] = 1
    first["slots"] = [{"d": "一", "p": 0}]
    second = deepcopy(first)
    second.update({"id": "grade-1-b", "grp": "一年級B組"})
    payload["resGroups"].append(second)

    data = engine.load_frontend_data(payload)

    with pytest.raises(ValueError, match="資源教師在週一早自修被兩個資源班分組重複固定"):
        engine.solve(data, time_limit=5, auto_schedule_tutor=False)


def test_resource_early_study_rejects_source_class_double_booking():
    payload = _resource_frontend_payload()
    first = payload["resGroups"][0]
    first["n"] = 1
    first["slots"] = [{"d": "一", "p": 0}]
    second = deepcopy(first)
    second.update({"id": "grade-1-b", "grp": "一年級B組", "t": "另一位資源教師"})
    payload["roster"]["另一位資源教師"] = "資源班教師"
    payload["resGroups"].append(second)

    data = engine.load_frontend_data(payload)

    with pytest.raises(ValueError, match="1甲在週一早自修被兩個資源班分組重複抽離"):
        engine.solve(data, time_limit=5, auto_schedule_tutor=False)


def test_per_class_arrangement_mode_controls_cp_sat_ownership():
    slots = [[1, 1, 1, 1, 0, 0, 0] for _ in range(5)]
    payload = {
        "classes": [{"g": 1, "i": 1, "code": "1甲", "tutor": "王老師"}],
        "roster": {"王老師": "導師"}, "rooms": {"R00": 99},
        "subjects": {"國語文": {"hours": [1, 0, 0, 0, 0, 0], "room": "R00",
                                   "banned": [], "block": "", "self": True, "pairMode": ""}},
        "gslot": {str(grade): slots for grade in range(1, 7)},
        "assign": {"1甲": {"國語文": "王老師"}},
        "override": {}, "locks": [], "blocked": [], "resGroups": [],
    }

    engine_payload = deepcopy(payload)
    engine_payload["assignmentModes"] = {"1甲": {"國語文": "engine"}}
    engine_data = engine.load_frontend_data(engine_payload)
    engine_schedule, *_ = engine.solve(engine_data, time_limit=5, auto_schedule_tutor=False)
    assert engine_data["assignment_modes"][("1甲", "國語文")] == "engine"
    assert len(engine_schedule) == 1

    tutor_payload = deepcopy(payload)
    tutor_payload["assignmentModes"] = {"1甲": {"國語文": "tutor"}}
    tutor_data = engine.load_frontend_data(tutor_payload)
    tutor_schedule, *_ = engine.solve(tutor_data, time_limit=5, auto_schedule_tutor=False)
    assert tutor_schedule == {}
    assert tutor_data["pool"]["1甲"] == [("國語文", 1, "王老師")]


def _native_frontend_payload():
    slots = [[1, 1, 1, 1, 0, 0, 0] for _ in range(5)]
    return {
        "classes": [
            {"g": 1, "i": 1, "code": "1甲", "tutor": "王導師"},
            {"g": 1, "i": 2, "code": "1乙", "tutor": "李導師"},
        ],
        "roster": {"王導師": "導師", "李導師": "導師", "直播教師": "科任",
                   "協同教師": "科任", "客語教師": "科任"},
        "rooms": {"R00": 99, "電腦教室": 1},
        "subjects": {"本土語文": {
            "hours": [1, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
            "block": "", "self": False, "pairMode": "",
        }},
        "gslot": {str(grade): slots for grade in range(1, 7)},
        "assign": {"1甲": {"本土語文": "舊配課教師"}},
        "override": {}, "locks": [], "blocked": [], "resGroups": [],
        "nativeLockEnabled": True,
        "nativeBands": [{"g": 1, "d": "二", "p": 1}],
        "nativeGroups": [{
            "g": 1, "d": "二", "p": 1, "lang": "原民語(直播)",
            "grp": "一年級原民語組", "sources": ["1甲", "1乙"], "students": 5,
            "mode": "直播共學", "t": "直播教師", "room": "電腦教室",
            "assistant": "協同教師",
        }],
    }


def _distributed_tutor_pull_payload(self_arrange=True):
    payload = _native_frontend_payload()
    payload["nativeArrangement"] = "distributed"
    payload["nativeBands"] = []
    payload["subjects"]["本土語文"]["hours"] = [0, 0, 0, 0, 0, 0]
    payload["subjects"]["綜合活動"] = {
        "hours": [1, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
        "block": "", "self": self_arrange, "pairMode": "",
    }
    payload["assign"] = {
        "1甲": {"綜合活動": "王導師"},
        "1乙": {"綜合活動": "李導師"},
    }
    payload["nativeGroups"] = [{
        "g": 1, "d": "二", "p": 1, "lang": "客語",
        "grp": "一年級客語組", "sources": ["1甲"],
        "students": 2, "mode": "實體", "arrangement": "distributed",
        "pullSubjects": ["綜合活動"],
        "t": "客語教師", "room": "R00", "assistant": "",
    }]
    return payload


def test_frontend_native_language_locks_classes_staff_and_room():
    data = engine.load_frontend_data(_native_frontend_payload())
    schedule, tasks, _, meta, overlay = engine.solve(data, time_limit=5)

    native_locks = [lock for lock in data["locks"] if lock["subj"] == "本土語文"]
    assert {(lock["class"], lock["day"], lock["p"]) for lock in native_locks} == {
        ("1甲", "二", 1), ("1乙", "二", 1)}
    assert ("直播教師", "二", 1) in data["teacher_limit"]
    assert ("協同教師", "二", 1) in data["teacher_limit"]
    assert ("電腦教室", "二", 1) in data["room_blocked"]
    assert data["native_groups"][0]["lang"] == "原民語(直播)"
    assert data["native_groups"][0]["sources"] == ["1甲", "1乙"]
    assert data["native_bands"][0]["sources"] == ["1甲", "1乙"]
    assert data["native_groups"][0]["students"] == 5
    assert data["teacher_weekly_load"]["直播教師"] == 1
    assert data["teacher_weekly_load"]["協同教師"] == 1
    assert data["teacher_weekly_load"]["舊配課教師"] == 1
    assert tasks[("1甲", "本土語文")]["t"] == "舊配課教師"
    assert schedule[("1甲", "二", 1)][1] == "舊配課教師"
    assert {(code, day, period) for code, day, period in schedule} == {
        ("1甲", "二", 1), ("1乙", "二", 1)}
    assert meta["completion"] == "complete"
    assert engine.validate(data, schedule, tasks, overlay) == []


def test_frontend_native_language_supports_three_classes_split_into_two_groups():
    payload = _native_frontend_payload()
    payload["classes"].append({"g": 1, "i": 3, "code": "1丙", "tutor": "丙班導師"})
    payload["roster"].update({"丙班導師": "導師", "閩南教師甲": "教支人員",
                              "閩南教師乙": "教支人員"})
    payload["nativeBands"][0]["sources"] = ["1甲", "1乙", "1丙"]
    payload["nativeGroups"] = [
        {
            "g": 1, "d": "二", "p": 1, "lang": "閩南語",
            "grp": "一年級閩南語A組", "sources": ["1甲", "1乙", "1丙"],
            "students": 28, "mode": "實體", "t": "閩南教師甲",
            "room": "R00", "assistant": "",
        },
        {
            "g": 1, "d": "二", "p": 1, "lang": "閩南語",
            "grp": "一年級閩南語B組", "sources": ["1甲", "1乙", "1丙"],
            "students": 27, "mode": "實體", "t": "閩南教師乙",
            "room": "R00", "assistant": "",
        },
    ]

    data = engine.load_frontend_data(payload)
    schedule, tasks, _, meta, overlay = engine.solve(data, time_limit=5)

    assert {(lock["class"], lock["day"], lock["p"]) for lock in data["locks"]
            if lock["subj"] == "本土語文"} == {
        ("1甲", "二", 1), ("1乙", "二", 1), ("1丙", "二", 1)}
    assert data["teacher_weekly_load"]["閩南教師甲"] == 1
    assert data["teacher_weekly_load"]["閩南教師乙"] == 1
    assert "舊配課教師" not in data["teacher_weekly_load"]
    assert all(tasks[(code, "本土語文")]["t"] == "" for code in ("1甲", "1乙", "1丙"))
    assert {(code, day, period) for code, day, period in schedule} == {
        ("1甲", "二", 1), ("1乙", "二", 1), ("1丙", "二", 1)}
    assert meta["completion"] == "complete"
    assert engine.validate(data, schedule, tasks, overlay) == []


def test_frontend_native_language_supports_distributed_groups_across_slots():
    payload = _native_frontend_payload()
    payload["nativeArrangement"] = "distributed"
    payload["nativeBands"] = []
    payload["subjects"]["本土語文"]["hours"] = [0, 0, 0, 0, 0, 0]
    payload["subjects"]["國語文"] = {
        "hours": [1, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
        "block": "", "self": False, "pairMode": "",
    }
    payload["assign"] = {
        "1甲": {"國語文": "王導師"},
        "1乙": {"國語文": "李導師"},
    }
    payload["nativeGroups"] = [
        {
            "g": 1, "d": "二", "p": 1, "lang": "客語",
            "grp": "一年級客語組", "sources": ["1甲", "1乙"],
            "students": 4, "mode": "實體", "arrangement": "distributed",
            "pullSubjects": ["國語文"],
            "t": "客語教師", "room": "R00", "assistant": "",
        },
        {
            "g": 1, "d": "三", "p": 2, "lang": "原住民族語",
            "grp": "一年級原住民族語組", "sources": ["1甲"],
            "students": 1, "mode": "直播共學", "arrangement": "distributed",
            "pullSubjects": ["國語文"],
            "t": "直播教師", "room": "電腦教室", "assistant": "協同教師",
        },
    ]

    data = engine.load_frontend_data(payload)

    assert data["native_bands"] == []
    assert data["native_external_sources"] == {"1甲", "1乙"}
    assert not any(lock["subj"] == "本土語文" for lock in data["locks"])
    assert ("客語教師", "二", 1) in data["teacher_limit"]
    assert ("直播教師", "三", 2) in data["teacher_limit"]
    assert ("協同教師", "三", 2) in data["teacher_limit"]
    assert ("電腦教室", "三", 2) in data["room_blocked"]
    assert data["teacher_weekly_load"]["客語教師"] == 1
    assert data["teacher_weekly_load"]["直播教師"] == 1
    assert data["teacher_weekly_load"]["協同教師"] == 1
    assert "舊配課教師" not in data["teacher_weekly_load"]


def test_distributed_native_pull_takes_over_tutor_arranged_subject():
    data = engine.load_frontend_data(_distributed_tutor_pull_payload())

    schedule, tasks, warnings, _, overlay = engine.solve(
        data, time_limit=5, auto_schedule_tutor=False)

    assert ("1甲", "綜合活動") in tasks
    assert schedule[("1甲", "二", 1)][0] == "綜合活動"
    assert all(row[0] != "綜合活動" for row in data["pool"].get("1甲", []))
    assert "引擎接手導師科目：1甲 綜合活動（語言抽離綁課）" in warnings
    assert engine.validate(data, schedule, tasks, overlay) == []


def test_h18_diagnoses_fixed_course_subject_conflict():
    payload = _distributed_tutor_pull_payload()
    payload["subjects"]["數學"] = {
        "hours": [1, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
        "block": "", "self": False, "pairMode": "",
    }
    payload["assign"]["1甲"]["數學"] = "王導師"
    payload["assign"]["1乙"]["數學"] = "李導師"
    payload["locks"] = [{"c": "1甲", "d": "二", "p": 1, "s": "數學"}]
    data = engine.load_frontend_data(payload)

    with pytest.raises(engine.InfeasibleScheduleError) as exc:
        engine.solve(data, time_limit=5, auto_schedule_tutor=False)

    assert any(
        item["title"] == "1甲 語言抽離與固定課衝突"
        and "綜合活動" in item["detail"] and "數學" in item["detail"]
        for item in exc.value.diagnostics)


def test_h18_diagnoses_resource_pull_subject_conflict():
    payload = _distributed_tutor_pull_payload(self_arrange=False)
    payload["subjects"]["數學"] = {
        "hours": [1, 0, 0, 0, 0, 0], "room": "R00", "banned": [],
        "block": "", "self": False, "pairMode": "",
    }
    payload["assign"]["1甲"]["數學"] = "王導師"
    payload["assign"]["1乙"]["數學"] = "李導師"
    payload["resGroups"] = [{
        "id": "resource-a", "grp": "一年級資源A",
        "sources": ["1甲"], "subj": "數學", "pullSubjects": ["數學"],
        "t": "協同教師", "n": 1, "scheduleMode": "fixed",
        "slots": [{"d": "二", "p": 1}],
    }]
    data = engine.load_frontend_data(payload)

    with pytest.raises(engine.InfeasibleScheduleError) as exc:
        engine.solve(data, time_limit=5, auto_schedule_tutor=False)

    assert any(
        item["title"] == "1甲 語言抽離與資源班抽離科目衝突"
        and "綜合活動" in item["detail"] and "數學" in item["detail"]
        for item in exc.value.diagnostics)


def test_frontend_distributed_native_language_does_not_require_class_subject():
    payload = _native_frontend_payload()
    payload["nativeArrangement"] = "distributed"
    payload["nativeBands"] = []
    payload["subjects"] = {
        "國語文": {
            "hours": [1, 0, 0, 0, 0, 0], "room": "R00",
            "banned": [], "block": "", "self": False, "pairMode": "",
        },
    }
    payload["assign"] = {
        "1甲": {"國語文": "王導師"},
        "1乙": {"國語文": "李導師"},
    }
    payload["nativeGroups"] = [
        {
            "g": 1, "d": "二", "p": 1, "lang": "客語",
            "grp": "一年級客語組", "sources": ["1甲"],
            "students": 2, "mode": "實體", "arrangement": "distributed",
            "pullSubjects": ["國語文"],
            "t": "客語教師", "room": "R00", "assistant": "",
        },
        {
            "g": 1, "d": "二", "p": 1, "lang": "原住民族語",
            "grp": "一年級原語組", "sources": ["1甲"],
            "students": 1, "mode": "實體", "arrangement": "distributed",
            "pullSubjects": ["國語文"],
            "t": "直播教師", "room": "電腦教室", "assistant": "",
        },
    ]

    data = engine.load_frontend_data(payload)

    assert data["native_external_sources"] == {"1甲"}
    assert data["teacher_weekly_load"]["客語教師"] == 1
    assert data["teacher_weekly_load"]["直播教師"] == 1
    assert not any(lock["subj"] == "本土語文" for lock in data["locks"])


def test_distributed_native_groups_allow_same_class_same_slot_with_shared_pull_subject():
    payload = _native_frontend_payload()
    payload["nativeArrangement"] = "distributed"
    payload["nativeBands"] = []
    payload["subjects"] = {
        "國語文": {
            "hours": [1, 0, 0, 0, 0, 0], "room": "R00",
            "banned": [], "block": "", "self": False, "pairMode": "",
        },
    }
    payload["assign"] = {
        "1甲": {"國語文": "王導師"},
        "1乙": {"國語文": "李導師"},
    }
    payload["nativeGroups"] = [
        {
            "g": 1, "d": "二", "p": 1, "lang": "客語", "grp": "客語組",
            "sources": ["1甲"], "students": 2, "arrangement": "distributed",
            "pullSubjects": ["國語文"], "t": "客語教師", "room": "R00",
        },
        {
            "g": 1, "d": "二", "p": 1, "lang": "原住民族語", "grp": "原語組",
            "sources": ["1甲"], "students": 1, "arrangement": "distributed",
            "pullSubjects": ["國語文"], "t": "直播教師", "room": "電腦教室",
        },
    ]

    data = engine.load_frontend_data(payload)
    schedule, tasks, *_ = engine.solve(data, time_limit=5)

    assert schedule[("1甲", "二", 1)][0] == "國語文"
    assert tasks[("1甲", "國語文")]["t"] == "王導師"


def test_distributed_native_groups_reject_disjoint_pull_subjects_same_class_slot():
    payload = _native_frontend_payload()
    payload["nativeArrangement"] = "distributed"
    payload["nativeBands"] = []
    payload["subjects"].update({
        "國語文": {
            "hours": [1, 0, 0, 0, 0, 0], "room": "R00",
            "banned": [], "block": "", "self": False, "pairMode": "",
        },
        "數學": {
            "hours": [1, 0, 0, 0, 0, 0], "room": "R00",
            "banned": [], "block": "", "self": False, "pairMode": "",
        },
    })
    payload["nativeGroups"] = [
        {
            "g": 1, "d": "二", "p": 1, "lang": "客語", "grp": "客語組",
            "sources": ["1甲"], "students": 2, "arrangement": "distributed",
            "pullSubjects": ["國語文"], "t": "客語教師", "room": "R00",
        },
        {
            "g": 1, "d": "二", "p": 1, "lang": "原住民族語", "grp": "原語組",
            "sources": ["1甲"], "students": 1, "arrangement": "distributed",
            "pullSubjects": ["數學"], "t": "直播教師", "room": "電腦教室",
        },
    ]

    with pytest.raises(ValueError, match="沒有共同可抽離科目"):
        engine.load_frontend_data(payload)


def test_frontend_distributed_native_language_rejects_stale_class_lock():
    payload = _native_frontend_payload()
    payload["nativeArrangement"] = "distributed"
    payload["nativeBands"] = []
    payload["nativeGroups"][0]["arrangement"] = "distributed"
    payload["locks"] = [{"c": "1甲", "d": "二", "p": 1, "s": "本土語文"}]

    with pytest.raises(ValueError, match="不應鎖住整班本土語文"):
        engine.load_frontend_data(payload)


def test_frontend_native_language_locks_only_selected_band_sources():
    payload = _native_frontend_payload()
    payload["nativeBands"][0]["sources"] = ["1甲"]
    payload["nativeGroups"][0]["sources"] = ["1甲"]

    data = engine.load_frontend_data(payload)

    assert data["native_bands"][0]["sources"] == ["1甲"]
    assert {(lock["class"], lock["day"], lock["p"]) for lock in data["locks"]
            if lock["subj"] == "本土語文"} == {("1甲", "二", 1)}


def test_frontend_native_language_group_sources_must_belong_to_band():
    payload = _native_frontend_payload()
    payload["nativeBands"][0]["sources"] = ["1甲"]
    payload["nativeGroups"][0]["sources"] = ["1甲", "1乙"]

    with pytest.raises(ValueError, match="未納入1年級共同課程"):
        engine.load_frontend_data(payload)


def test_frontend_native_language_allows_band_without_extraction_group():
    payload = _native_frontend_payload()
    payload["nativeGroups"] = []
    payload["assign"]["1乙"] = {"本土語文": "李導師"}

    data = engine.load_frontend_data(payload)

    assert data["native_groups"] == []
    assert data["teacher_weekly_load"]["舊配課教師"] == 1
    assert {(lock["class"], lock["day"], lock["p"]) for lock in data["locks"]} == {
        ("1甲", "二", 1), ("1乙", "二", 1)}


def test_frontend_native_language_rejects_uncovered_band_without_original_teacher():
    payload = _native_frontend_payload()
    payload["nativeGroups"] = []

    with pytest.raises(ValueError, match="1乙.*沒有平行分組.*未指定原班授課教師"):
        engine.load_frontend_data(payload)


def test_legacy_minnan_group_keeps_group_teacher_and_suppresses_base_assignment():
    payload = _native_frontend_payload()
    payload["nativeGroups"][0]["lang"] = "閩南語"

    data = engine.load_frontend_data(payload)
    schedule, tasks, warnings, *_ = engine.solve(data, time_limit=5)

    assert "舊配課教師" not in data["teacher_weekly_load"]
    assert tasks[("1甲", "本土語文")]["t"] == ""
    assert schedule[("1甲", "二", 1)][1] == ""
    assert not any("鎖定課未配教師" in warning for warning in warnings)


def test_frontend_limits_use_exact_class_grade_and_roster_targets():
    payload = _native_frontend_payload()
    payload["nativeLockEnabled"] = False
    payload["classes"][1]["code"] = "1壬"
    payload["roster"]["丁美玲"] = "科任"
    payload["limits"] = [
        ["1壬", "一", "1", "不可排", ""],
        ["一年級", "二", "2", "不可排", ""],
        ["丁美玲", "三", "3", "不可排", ""],
    ]

    data = engine.load_frontend_data(payload)

    assert ("1壬", "一", 1) in data["class_limit"]
    assert (1, "二", 2) in data["grade_limit"]
    assert ("丁美玲", "三", 3) in data["teacher_limit"]


def test_frontend_limits_reject_unknown_target():
    payload = _native_frontend_payload()
    payload["nativeLockEnabled"] = False
    payload["limits"] = [["打錯教師", "一", "1", "不可排", ""]]

    with pytest.raises(ValueError, match="不排課時間引用不存在"):
        engine.load_frontend_data(payload)


def test_frontend_fixed_course_reports_unavailable_grade_slot():
    payload = _native_frontend_payload()
    payload["nativeLockEnabled"] = False
    payload["locks"] = [{"c": "1甲", "d": "一", "p": 5, "s": "本土語文"}]

    with pytest.raises(ValueError, match="固定課不在該年級可排時段"):
        engine.load_frontend_data(payload)


def test_frontend_fixed_course_reports_teacher_collision():
    payload = _native_frontend_payload()
    payload["nativeLockEnabled"] = False
    payload["assign"]["1乙"] = {"本土語文": "舊配課教師"}
    payload["locks"] = [
        {"c": "1甲", "d": "一", "p": 1, "s": "本土語文"},
        {"c": "1乙", "d": "一", "p": 1, "s": "本土語文"},
    ]

    with pytest.raises(ValueError, match="舊配課教師的固定課發生衝堂"):
        engine.load_frontend_data(payload)


def test_frontend_native_language_rejects_duplicate_staff_assignment():
    payload = _native_frontend_payload()
    payload["nativeGroups"].append({
        "g": 1, "d": "二", "p": 1, "lang": "客語", "grp": "一年級客語組",
        "sources": ["1甲"], "t": "直播教師", "room": "R00",
    })

    with pytest.raises(ValueError, match="重複指派本土語分組"):
        engine.load_frontend_data(payload)


def test_frontend_native_language_enforces_room_capacity():
    payload = _native_frontend_payload()
    payload["nativeGroups"].append({
        "g": 1, "d": "二", "p": 1, "lang": "客語", "grp": "一年級客語組",
        "sources": ["1甲"], "t": "客語教師", "room": "電腦教室",
    })

    with pytest.raises(ValueError, match="超過本土語分組可用容量"):
        engine.load_frontend_data(payload)


def test_frontend_native_language_rejects_unknown_source_class():
    payload = _native_frontend_payload()
    payload["nativeGroups"][0]["sources"] = ["1甲", "2甲"]

    with pytest.raises(ValueError, match="引用不存在的來源班級：2甲"):
        engine.load_frontend_data(payload)


def test_frontend_native_language_teacher_must_come_from_roster():
    payload = _native_frontend_payload()
    payload["nativeGroups"][0]["t"] = "名冊外教師"

    with pytest.raises(ValueError, match="授課教師不在教師名冊"):
        engine.load_frontend_data(payload)


def test_frontend_native_language_migrates_legacy_group_fields():
    payload = _native_frontend_payload()
    payload.pop("nativeBands")
    for key in ("grp", "sources", "students", "mode"):
        payload["nativeGroups"][0].pop(key)

    data = engine.load_frontend_data(payload)

    assert data["native_bands"] == [{
        "g": 1, "d": "二", "p": 1, "sources": ["1甲", "1乙"]}]
    assert data["native_groups"][0]["sources"] == ["1甲", "1乙"]
    assert data["native_groups"][0]["grp"].startswith("1年級原民語")


def test_frontend_native_language_lock_is_optional():
    payload = _native_frontend_payload()
    payload["nativeLockEnabled"] = False

    data = engine.load_frontend_data(payload)

    assert data["native_lock_enabled"] is False
    assert data["native_groups"] == []
    assert not any(lock["subj"] == "本土語文" for lock in data["locks"])


def test_v5_solves_and_passes_independent_validation():
    data = engine.load_data(V5_TEMPLATE)
    sched, tasks, warn, meta, overlay = engine.solve(data, time_limit=5)

    assert meta["status"] in {"OPTIMAL", "FEASIBLE"}
    assert sched
    assert data["pool"]
    assert meta["quality_penalty_total"] == pytest.approx(meta["penalty"])
    assert engine.validate(data, sched, tasks, overlay) == []


def test_v5_formal_mode_auto_schedules_tutor_pool_and_reports_completeness():
    data = engine.load_data(V5_TEMPLATE)
    sched, tasks, warn, meta, overlay = engine.solve(
        data, time_limit=5, auto_schedule_tutor=True)

    assert meta["status"] in {"OPTIMAL", "FEASIBLE"}
    assert meta["pool_total"] == 0
    assert meta["scheduled_total"] == len(sched)
    assert meta["required_total"] >= meta["scheduled_total"]
    assert meta["completion"] == "partial"
    assert meta["missing_total"] > 0
    assert meta["weekly_cap_violations"] == []
    assert any("尚未填寫教師每週基準節數" in item for item in meta["compliance_warnings"])
    assert engine.validate(data, sched, tasks, overlay) == []


def test_v4_parser_is_preserved():
    data = engine.load_data(V4_TEMPLATE)

    assert data.get("schema_version", 4) == 4
    assert data["classes"]
    assert data["subjects"]


def _validation_fixture(classes, subject_info):
    return {
        "classes": classes,
        "grade_slot": {(1, day, period): True for day in engine.DAYS for period in engine.PERIODS},
        "teacher_limit": set(),
        "grade_limit": set(),
        "class_limit": set(),
        "room_blocked": set(),
        "rooms": {"R00": 999},
        "locks": [],
        "roster": {},
    }, subject_info


def test_validator_detects_nonconsecutive_two_period_block():
    data, info = _validation_fixture(
        [{"code": "1甲", "grade": 1}],
        {"block": "2連堂", "banned": set()},
    )
    tasks = {("1甲", "視覺藝術"): {"h": 2, "info": info}}
    sched = {
        ("1甲", "一", 1): ("視覺藝術", "張老師", "R00"),
        ("1甲", "一", 3): ("視覺藝術", "張老師", "R00"),
    }

    assert any("H09違反" in error for error in engine.validate(data, sched, tasks))


def test_validator_detects_tutor_daily_hard_cap():
    classes = [{"code": f"1班{i}", "grade": 1} for i in range(1, 8)]
    data, info = _validation_fixture(classes, {"block": "", "banned": set()})
    data["roster"] = {"王導師": "導師"}
    tasks = {(c["code"], "課程"): {"h": 1, "info": info} for c in classes}
    sched = {(c["code"], "一", i): ("課程", "王導師", "R00")
             for i, c in enumerate(classes, 1)}

    assert any("每日上限違反" in error for error in engine.validate(data, sched, tasks))


def test_excel_safe_neutralizes_formula_prefixes():
    for value in ("=1+1", "+SUM(A1:A2)", "-2+3", "@SUM(A1:A2)", "  =CMD()"):
        protected = engine.excel_safe(value)
        assert protected.startswith("'")
        assert protected[1:] == value

    assert engine.excel_safe("一般文字") == "一般文字"
    assert engine.excel_safe(12) == 12


def test_infeasibility_diagnosis_finds_class_and_teacher_capacity():
    task = {"h": 3, "t": "王老師", "room": "R00",
            "info": {"block": "", "banned": set()}}
    data = {
        "locks": [], "rooms": {"R00": 999}, "room_names": {},
        "room_blocked": set(), "native_groups": [], "roster": {"王老師": "科任"},
    }
    tasks = {("1甲", "數學"): task}
    candidates = {
        ("1甲", "數學", "一", 1): object(),
        ("1甲", "數學", "二", 1): object(),
    }

    diagnostics = engine.diagnose_infeasibility(data, tasks, candidates)

    assert any(item["title"] == "1甲 可排節次不足" for item in diagnostics)
    assert any(item["title"] == "王老師的授課容量不足" for item in diagnostics)
    assert all(item["confirmed"] is True for item in diagnostics)
